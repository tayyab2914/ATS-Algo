
import pandas as pd
import numpy as np
import argparse
import json
import collections
from datetime import datetime, timezone
from dataclasses import dataclass
from typing import Optional, List, Dict, Any, Tuple
from pathlib import Path
from colorama import init
init()
import sys
sys.stdout.reconfigure(encoding="utf-8")

# Base window list — overridden at runtime by --selection_window
# Only windows <= selection_window are included in the output
WINDOWS = [30, 90, 180, 360]
LEVERAGES = [1, 2, 3, 4, 5, 6, 7, 10, 12, 15]
# -------------------------------
# Utilities
# -------------------------------

from pathlib import Path
import re

def _infer_token_and_tf_from_csv(csv_path: str):
    if not csv_path:
        return "UNKNOWN", None

    name = Path(csv_path).stem.upper()
    name = name.replace("-", "_")

    tf = None

    # TradingView .P<number> format
    m = re.search(r"\.P(\d+)$", name)
    if m:
        minutes = int(m.group(1))
        if minutes % 1440 == 0:
            tf = f"{minutes // 1440}d"
        elif minutes % 60 == 0:
            tf = f"{minutes // 60}h"
        else:
            tf = f"{minutes}m"
        name = name[:m.start()]

    # Standard _1m / _1h
    if tf is None:
        m = re.search(r"_(\d+)(M|H|D)$", name)
        if m:
            tf = f"{m.group(1)}{m.group(2).lower()}"
            name = name[:m.start()]

    parts = name.split("_")

    EXCHANGES = {
        "BINANCE", "OKX", "BYBIT", "BITGET",
        "KUCOIN", "COINBASE", "KRAKEN"
    }

    if parts and parts[0] in EXCHANGES:
        parts = parts[1:]

    token = parts[0]
    token = re.sub(r"(USDT|USD|USDC|PERP|SWAP)$", "", token)

    return token, tf

def _pick_row(dfX, sl, be, basis):
    if dfX is None:
        return None
    m = (
        (dfX["SL (%)"] == sl) &
        (dfX["BE (after TP #)"] == be) &
        (dfX["Basis"] == basis)
    )
    sub = dfX.loc[m]
    if len(sub) == 0:
        return None
    return sub.iloc[0]


# -------------------------------
# Utilities
# -------------------------------

def _sig_present(x) -> bool:
    if pd.isna(x):
        return False
    if isinstance(x, str):
        return x.strip() != ""
    return bool(x)

# -------------------------------
# Step 1: Build trades from signals
# -------------------------------

@dataclass
class Trade:
    side: str               # "long" | "short"
    entry_time: pd.Timestamp
    entry_price: float
    exit_time: pd.Timestamp
    exit_price: float
    bars_in_trade: int
    pnl_pct: float
    mfe_pct: float
    mae_pct: float
    exit_reason: str        # "exit_signal" | "new_entry"

def load_signals_csv(path: str) -> pd.DataFrame:
    df = pd.read_csv(path)
    # robust time parsing → UTC
    df["ts"] = pd.to_datetime(df["time"], utc=True, errors="coerce")
    df = df.sort_values("ts").reset_index(drop=True)

    # normalize signals (non-empty = True)
    df["buy_sig"] = df["Buy Entry Signal"].apply(_sig_present)
    df["sell_sig"] = df["Sell Entry Signal"].apply(_sig_present)
    df["exit_sig"] = df["Exit Signal"].apply(_sig_present)
    return df

def build_trades(df: pd.DataFrame) -> List[Trade]:
    """
    Build trades from signal dataframe with pyramiding:

    - Mehrere aufeinanderfolgende Signale in dieselbe Richtung (long/short)
    erzeugen mehrere, gleich gewichtete Trades (Pyramiding).
    - Alle offenen Trades einer Seite werden geschlossen
        * bei einem Exit-Signal, oder
        * bei einem Entry-Signal in die entgegengesetzte Richtung.
    - Jede Pyramiden-Position ist ein vollwertiger Trade mit eigener
    Entry time, entry price, MFE/MAE and PnL, which then runs normally
    through the TP/SL/BE simulation.
    """
    recs: List[Trade] = []
    n = len(df)

    # Liste offener Trades (Pyramiding)
    # Jeder Eintrag: {"side": "long"/"short", "entry_idx": int, "entry_price": float}
    open_trades: List[Dict[str, Any]] = []

    i = 0
    while i < n - 1:  # we always need the next bar for execution at open
        row = df.iloc[i]
        next_open = df.iloc[i + 1]["open"]

        buy_sig = bool(row["buy_sig"])
        sell_sig = bool(row["sell_sig"])
        exit_sig = bool(row["exit_sig"])

        # Which side does this bar signal (if any)?
        signal_side: Optional[str] = None
        if buy_sig:
            signal_side = "long"
        elif sell_sig:
            signal_side = "short"

        # Aktuell offene Seite (falls Trades offen sind)
        current_side: Optional[str] = open_trades[0]["side"] if open_trades else None

        # ----------------------------
        # 1) Muss der gesamte Stack geschlossen werden?
        #    a) Exit signal closes all open trades on this side.
        #    b) Signal in opposite direction closes all open trades
        #       on this side and optionally opens the counter position.
        # ----------------------------
        exit_trades = False
        exit_reason = None
        new_side_after_exit: Optional[str] = None

        if exit_sig and open_trades:
            # Explicit exit signal: close all open trades
            exit_trades = True
            exit_reason = "exit_signal"
            # Additionally a new entry signal can occur on the same bar
            new_side_after_exit = signal_side
        elif open_trades and signal_side is not None and signal_side != current_side:
            # Direction change: close all open trades, then open new side
            exit_trades = True
            exit_reason = "new_entry"
            new_side_after_exit = signal_side

        if exit_trades:
            if i + 1 < n:
                exit_idx = i + 1
                exit_time = df.iloc[exit_idx]["ts"]
                exit_price = float(df.iloc[exit_idx]["open"])

                for ot in open_trades:
                    e_idx = ot["entry_idx"]
                    entry_price = ot["entry_price"]
                    side = ot["side"]

                    # Fenster von Entry bis inklusive Exit-Kerze
                    window = df.iloc[e_idx: exit_idx + 1]
                    highs = window["high"].to_numpy()
                    lows = window["low"].to_numpy()

                    if side == "long":
                        mfe = (np.max(highs) / entry_price - 1) * 100
                        mae = (np.min(lows)  / entry_price - 1) * 100
                        pnl = (exit_price / entry_price - 1) * 100
                    else:
                        mfe = (entry_price / np.min(lows)  - 1) * 100  # favorable down move
                        mae = (entry_price / np.max(highs) - 1) * 100  # adverse up move (negativ)
                        pnl = (entry_price / exit_price - 1) * 100

                    recs.append(Trade(
                        side=side,
                        entry_time=df.iloc[e_idx]["ts"],
                        entry_price=entry_price,
                        exit_time=exit_time,
                        exit_price=exit_price,
                        bars_in_trade=int(exit_idx - e_idx),
                        pnl_pct=float(pnl),
                        mfe_pct=float(mfe),
                        mae_pct=float(mae),
                        exit_reason=exit_reason,
                    ))

            # Stack geleert
            open_trades = []

            # Optionally open new position on this bar (other side)
            if new_side_after_exit is not None and i + 1 < n:
                entry_idx = i + 1
                entry_price = float(df.iloc[entry_idx]["open"])
                open_trades.append({
                    "side": new_side_after_exit,
                    "entry_idx": entry_idx,
                    "entry_price": entry_price,
                })

            i += 1
            continue

        # ----------------------------
        # 2) No stack exit: optionally open new pyramid position
        # ----------------------------
        if signal_side is not None:
            if current_side is None:
                # No position open yet: open first position in this direction
                if i + 1 < n:
                    entry_idx = i + 1
                    entry_price = float(df.iloc[entry_idx]["open"])
                    open_trades.append({
                        "side": signal_side,
                        "entry_idx": entry_idx,
                        "entry_price": entry_price,
                    })
            else:
                if signal_side == current_side:
                    # Gleichgerichtetes Signal => Pyramiding: weiteres vollwertiges Leg
                    if i + 1 < n:
                        entry_idx = i + 1
                        entry_price = float(df.iloc[entry_idx]["open"])
                        open_trades.append({
                            "side": signal_side,
                            "entry_idx": entry_idx,
                            "entry_price": entry_price,
                        })
                else:
                    # Opposite direction with open trades would already be handled as exit_trades above
                    pass

        # Offene Trades bleiben bestehen, bis ein Exit-Signal oder Richtungswechsel kommt.
        i += 1

    # At end of data, open trades remain unchanged (same as original logic).
    return recs

# -------------------------------
# Step 2: TP/SL Simulator
# -------------------------------


def _baseline_outcome(window: pd.DataFrame, side: str, entry_price: float,
                    tp_pcts: List[float], tp_weights: List[float],
                    sl_pct: float) -> Dict[str, Any]:
    """Evaluate terminal outcome WITHOUT BE. Vectorized with NumPy arrays."""
    tp_hit = [False]*len(tp_pcts)
    remaining_weight = 1.0
    stop_price = entry_price * (1 - sl_pct/100.0) if side == "long" else entry_price * (1 + sl_pct/100.0)
    partial_events = 0

    highs = window["high"].to_numpy(dtype=float)
    lows  = window["low"].to_numpy(dtype=float)

    for i in range(len(highs)):
        high = highs[i]; low = lows[i]

        if side == "long":
            if low <= stop_price:
                return {"terminal":"SL", "partial_events":partial_events, "full_tp":False, "tp_hits": tp_hit}
            for idx, tp_pct in enumerate(tp_pcts):
                if not tp_hit[idx]:
                    tp_price = entry_price * (1 + tp_pct/100.0)
                    if high >= tp_price:
                        tp_hit[idx] = True
                        partial_events += 1
                        remaining_weight -= tp_weights[idx]
                        if remaining_weight <= 1e-9:
                            return {"terminal":"full_TP", "partial_events":partial_events, "full_tp":True, "tp_hits": tp_hit}
        else:
            if high >= stop_price:
                return {"terminal":"SL", "partial_events":partial_events, "full_tp":False, "tp_hits": tp_hit}
            for idx, tp_pct in enumerate(tp_pcts):
                if not tp_hit[idx]:
                    tp_price = entry_price * (1 - tp_pct/100.0)
                    if low <= tp_price:
                        tp_hit[idx] = True
                        partial_events += 1
                        remaining_weight -= tp_weights[idx]
                        if remaining_weight <= 1e-9:
                            return {"terminal":"full_TP", "partial_events":partial_events, "full_tp":True, "tp_hits": tp_hit}

    return {"terminal":"generic_exit", "partial_events":partial_events, "full_tp":False, "tp_hits": tp_hit}


def simulate_single_trade(window: pd.DataFrame, side: str, entry_price: float,
                        tp_pcts: List[float], tp_weights: List[float],
                        sl_pct: float, sl_to_be_trigger: Optional[int]) -> Dict[str, Any]:
    """Simulate a single trade bar-by-bar.
    Vectorized with NumPy arrays — avoids iterrows() overhead.
    Output is identical to the original implementation.
    """
    n_tps = len(tp_pcts)
    tp_hit = [False] * n_tps
    tp_exec_weights = [0.0] * n_tps

    remaining_weight = 1.0
    realized_pnl = 0.0
    breakdown = collections.Counter()
    close_reason: Optional[str] = None
    remainder_exit_weight: float = 0.0

    # Extract arrays once — avoids repeated .iloc[] overhead
    highs = window["high"].to_numpy(dtype=float)
    lows  = window["low"].to_numpy(dtype=float)
    opens = window["open"].to_numpy(dtype=float)
    n_bars = len(highs)

    if n_bars == 0:
        return {
            "realized_pnl_pct": 0.0,
            "breakdown": {},
            "tp_exec_weights": [0.0] * len(tp_pcts),
            "remaining_weight": 1.0,
            "remainder_exit_weight": 0.0,
            "close_reason": "generic_exit",
        }

    stop_price = entry_price * (1 - sl_pct/100.0) if side == "long" else entry_price * (1 + sl_pct/100.0)
    be_pending = False

    for i in range(n_bars):
        if be_pending:
            stop_price = entry_price
            be_pending = False

        high = highs[i]; low = lows[i]
        is_last_bar = (i == n_bars - 1)

        if side == "long":
            # On the last bar the exit signal fires at open — check TP hits
            # at the open price first, then only check SL/BE on intrabar wick
            # if it's NOT the last bar. This prevents a wick to BE on the exit
            # bar from overriding an actual exit at a profitable open price.
            if not is_last_bar and low <= stop_price:
                remainder_exit_weight = remaining_weight
                if abs(stop_price - entry_price) < 1e-12:
                    breakdown["BE"] += 1
                    close_reason = "BE"
                else:
                    breakdown["SL"] += 1
                    close_reason = "SL"
                realized_pnl += remaining_weight * ((stop_price/entry_price - 1) * 100)
                remaining_weight = 0.0
                break

            for idx in range(n_tps):
                if not tp_hit[idx]:
                    tp_price = entry_price * (1 + tp_pcts[idx]/100.0)
                    if high >= tp_price:
                        w = tp_weights[idx]
                        realized_pnl += w * ((tp_price/entry_price - 1) * 100)
                        remaining_weight -= w
                        tp_hit[idx] = True
                        tp_exec_weights[idx] = float(w)
                        breakdown["partial_TP"] += 1
                        if sl_to_be_trigger is not None and (idx+1) == sl_to_be_trigger:
                            be_pending = True
            if remaining_weight <= 1e-9:
                breakdown["full_TP"] += 1
                close_reason = "full_TP"
                remainder_exit_weight = 0.0
                break

        else:  # short
            if not is_last_bar and high >= stop_price:
                remainder_exit_weight = remaining_weight
                if abs(stop_price - entry_price) < 1e-12:
                    breakdown["BE"] += 1
                    close_reason = "BE"
                else:
                    breakdown["SL"] += 1
                    close_reason = "SL"
                realized_pnl += remaining_weight * ((entry_price/stop_price - 1) * 100)
                remaining_weight = 0.0
                break

            for idx in range(n_tps):
                if not tp_hit[idx]:
                    tp_price = entry_price * (1 - tp_pcts[idx]/100.0)
                    if low <= tp_price:
                        w = tp_weights[idx]
                        realized_pnl += w * ((entry_price/tp_price - 1) * 100)
                        remaining_weight -= w
                        tp_hit[idx] = True
                        tp_exec_weights[idx] = float(w)
                        breakdown["partial_TP"] += 1
                        if sl_to_be_trigger is not None and (idx+1) == sl_to_be_trigger:
                            be_pending = True
            if remaining_weight <= 1e-9:
                breakdown["full_TP"] += 1
                close_reason = "full_TP"
                remainder_exit_weight = 0.0
                break

    if remaining_weight > 1e-9:
        last_open = opens[-1]
        remainder_exit_weight = remaining_weight
        if side == "long":
            realized_pnl += remaining_weight * ((last_open/entry_price - 1) * 100)
        else:
            realized_pnl += remaining_weight * ((entry_price/last_open - 1) * 100)
        breakdown["generic_exit"] += 1
        close_reason = "generic_exit"

    return {
        "realized_pnl_pct": float(realized_pnl),
        "breakdown": dict(breakdown),
        "tp_exec_weights": tp_exec_weights,
        "remaining_weight": float(remaining_weight),
        "remainder_exit_weight": float(remainder_exit_weight),
        "close_reason": close_reason,
    }



def simulate_single_trade_with_meta(window: pd.DataFrame, side: str, entry_price: float,
                                    tp_pcts: List[float], tp_weights: List[float],
                                    sl_pct: float, sl_to_be_trigger: Optional[int]) -> Dict[str, Any]:
    """Same core logic as simulate_single_trade(), but returns meta for fee modelling.
    Vectorized with NumPy arrays — avoids iterrows() overhead.
    """
    import collections

    n_tps = len(tp_pcts)
    tp_hit = [False] * n_tps
    tp_exec_weights = [0.0] * n_tps
    remaining_weight = 1.0
    realized_pnl = 0.0

    stop_price = entry_price * (1 - sl_pct/100.0) if side == "long" else entry_price * (1 + sl_pct/100.0)
    be_pending = False
    breakdown = collections.Counter()
    close_reason = None
    remainder_exit_weight = 0.0

    # Extract arrays once
    highs = window["high"].to_numpy(dtype=float)
    lows  = window["low"].to_numpy(dtype=float)
    opens = window["open"].to_numpy(dtype=float)
    n_bars = len(highs)

    if n_bars == 0:
        return {
            "realized_pnl_pct": 0.0,
            "breakdown": {},
            "tp_exec_weights": [0.0] * len(tp_pcts),
            "remainder_exit_weight": 0.0,
            "close_reason": "generic_exit",
        }

    for i in range(n_bars):
        if be_pending:
            stop_price = entry_price
            be_pending = False

        high = highs[i]; low = lows[i]
        is_last_bar = (i == n_bars - 1)

        if side == "long":
            if not is_last_bar and low <= stop_price:
                remainder_exit_weight = remaining_weight
                if abs(stop_price - entry_price) < 1e-12:
                    breakdown["BE"] += 1
                    close_reason = "BE"
                else:
                    breakdown["SL"] += 1
                    close_reason = "SL"
                realized_pnl += remaining_weight * ((stop_price/entry_price - 1) * 100)
                remaining_weight = 0.0
                break

            for idx in range(n_tps):
                if not tp_hit[idx]:
                    tp_price = entry_price * (1 + tp_pcts[idx]/100.0)
                    if high >= tp_price:
                        w = tp_weights[idx]
                        realized_pnl += w * ((tp_price/entry_price - 1) * 100)
                        remaining_weight -= w
                        tp_hit[idx] = True
                        tp_exec_weights[idx] = float(w)
                        breakdown["partial_TP"] += 1
                        if sl_to_be_trigger is not None and (idx+1) == sl_to_be_trigger:
                            be_pending = True
            if remaining_weight <= 1e-9:
                breakdown["full_TP"] += 1
                close_reason = "full_TP"
                break

        else:
            if not is_last_bar and high >= stop_price:
                remainder_exit_weight = remaining_weight
                if abs(stop_price - entry_price) < 1e-12:
                    breakdown["BE"] += 1
                    close_reason = "BE"
                else:
                    breakdown["SL"] += 1
                    close_reason = "SL"
                realized_pnl += remaining_weight * ((entry_price/stop_price - 1) * 100)
                remaining_weight = 0.0
                break

            for idx in range(n_tps):
                if not tp_hit[idx]:
                    tp_price = entry_price * (1 - tp_pcts[idx]/100.0)
                    if low <= tp_price:
                        w = tp_weights[idx]
                        realized_pnl += w * ((entry_price/tp_price - 1) * 100)
                        remaining_weight -= w
                        tp_hit[idx] = True
                        tp_exec_weights[idx] = float(w)
                        breakdown["partial_TP"] += 1
                        if sl_to_be_trigger is not None and (idx+1) == sl_to_be_trigger:
                            be_pending = True
            if remaining_weight <= 1e-9:
                breakdown["full_TP"] += 1
                close_reason = "full_TP"
                break

    if remaining_weight > 1e-9:
        last_open = opens[-1]
        remainder_exit_weight = remaining_weight
        if side == "long":
            realized_pnl += remaining_weight * ((last_open/entry_price - 1) * 100)
        else:
            realized_pnl += remaining_weight * ((entry_price/last_open - 1) * 100)
        breakdown["generic_exit"] += 1
        close_reason = "generic_exit"

    return {
        "realized_pnl_pct": float(realized_pnl),
        "breakdown": dict(breakdown),
        "tp_exec_weights": tp_exec_weights,
        "remainder_exit_weight": float(remainder_exit_weight),
        "close_reason": close_reason,
    }


def evaluate_scenario_net(trades: List[Trade], df: pd.DataFrame,
                        tp_pcts: List[float], tp_weights: List[float],
                        sl_pct: float, sl_to_be_trigger: Optional[int],
                        initial_dd_hours: float,
                        maker_fee_pct: float,
                        taker_fee_pct: float,):
                        
    """Compute NET metrics — identical counting logic to GROSS, only PnL differs by fees.
    Winrate, TP counts, terminal counts, avg win/loss are all fee-independent and must
    match GROSS exactly. Only compounded profit, max DD, fixed-size, and profit factor differ.
    """
    import numpy as np
    import collections

    pnls_net: list[float] = []
    terminal_counts: collections.Counter[str] = collections.Counter()
    tp_reached_counts = [0]*len(tp_pcts)
    be_after_tp_direct = {1: 0, 2: 0, 3: 0}

    fees_list: list[float] = []
    initial_dd_values: list[float] = []

    for tr in trades:
        window = df[(df["ts"] >= tr.entry_time) & (df["ts"] <= tr.exit_time)]
        if window.empty:
            continue

        # Use same simulate function as GROSS for consistent counting
        sim = simulate_single_trade_with_meta(window, tr.side, tr.entry_price, tp_pcts, tp_weights, sl_pct, sl_to_be_trigger)
        gross_pnl = float(sim["realized_pnl_pct"])

        tp_exec = sim.get("tp_exec_weights", [0.0]*len(tp_pcts))

        # TP reach counts — use _baseline_outcome same as GROSS
        # so counts are identical regardless of fees
        base = _baseline_outcome(window, tr.side, tr.entry_price, tp_pcts, tp_weights, sl_pct)
        for i, hit in enumerate(base["tp_hits"]):
            if i < len(tp_reached_counts) and hit:
                tp_reached_counts[i] += 1

        # Terminal type — count PER TRADE same as GROSS (not as events)
        br = sim.get("breakdown", {})
        if br.get("SL", 0) > 0:
            terminal = "SL"
        elif br.get("BE", 0) > 0:
            terminal = "BE"
        elif br.get("full_TP", 0) > 0:
            terminal = "full_TP"
        elif br.get("generic_exit", 0) > 0:
            terminal = "generic_exit"
        elif br.get("partial_TP", 0) > 0:
            terminal = "partial_TP"
        else:
            terminal = "unknown"
        terminal_counts[terminal] += 1

        maker_notional = float(sum(tp_exec))
        rem = float(sim.get("remainder_exit_weight", 0.0))
        close_reason = sim.get("close_reason", "generic_exit")

        if close_reason == "BE":
            hits = [i+1 for i,w in enumerate(tp_exec) if w is not None and float(w) > 0.0]
            if hits:
                h = max(hits)
                if h in be_after_tp_direct:
                    be_after_tp_direct[h] += 1

        # Fees: entry taker on full notional
        entry_fee = float(taker_fee_pct)

        # Remainder exit fee by terminal type
        if close_reason == "BE":
            remainder_fee_rate = float(maker_fee_pct)
        elif close_reason in ("SL", "generic_exit"):
            remainder_fee_rate = float(taker_fee_pct)
        else:
            remainder_fee_rate = 0.0

        exit_fee = maker_notional*float(maker_fee_pct) + rem*remainder_fee_rate
        fees_notional = entry_fee + exit_fee
        fees_list.append(fees_notional)

        net_pnl = gross_pnl - fees_notional
        pnls_net.append(net_pnl)

        # Initial DD (same as original definition)
        if initial_dd_hours and initial_dd_hours > 0:
            cutoff = tr.entry_time + pd.Timedelta(hours=float(initial_dd_hours))
            wdd = df[(df["ts"] >= tr.entry_time) & (df["ts"] <= cutoff)]
            if not wdd.empty:
                entry_px = tr.entry_price
                if tr.side == "long":
                    dd = max(0.0, (entry_px - float(wdd["low"].min()))/entry_px * 100.0)
                else:
                    dd = max(0.0, (float(wdd["high"].max()) - entry_px)/entry_px * 100.0)
                initial_dd_values.append(dd)

    pnls_net = np.array(pnls_net, dtype=float)
    n_trades = int(len(pnls_net))
    if n_trades == 0:
        return {
            "n_trades": 0,
            "winrate": 0.0,
            "avg_win_pct": 0.0,
            "avg_loss_pct": 0.0,
            "profit_factor": float("nan"),
            "counts": {},
            "avg_fees_notional_pct": 0.0,
            "total_fees_notional_pct": 0.0,
        }

    wins = pnls_net[pnls_net > 0]
    losses = pnls_net[pnls_net < 0]
    winrate = len(wins)/n_trades if n_trades else 0.0
    avg_win = float(np.mean(wins)) if len(wins) else 0.0
    avg_loss = float(np.mean(losses)) if len(losses) else 0.0
    profit_factor = float(np.sum(wins)/abs(np.sum(losses))) if len(losses) else float("inf")

    stats1 = _equity_stats_from_returns(pnls_net)
    compounded_net_profit_pct = stats1["compounded_pct"]
    max_dd = stats1["max_dd_pct"]

    lev_stats = {}
    fixed_totals = {}

    for L in LEVERAGES:
        rL = np.maximum(pnls_net * float(L), -100.0)
        st = _equity_stats_from_returns(rL)
        lev_stats[f"L{L}x_compounded_pct"] = st["compounded_pct"]
        lev_stats[f"L{L}x_max_drawdown_pct"] = st["max_dd_pct"]
        fixed_totals[f"FSum_L{L}x_pct"] = float(np.sum(rL))

    counts = {
        "Closed: SL": int(terminal_counts.get("SL", 0)),
        "Closed: BE": int(terminal_counts.get("BE", 0)),
        "Closed: partial TP": int(terminal_counts.get("partial_TP", 0)),
        "Closed: full TP": int(terminal_counts.get("full_TP", 0)),
        "Closed: generic exit (pure)": int(terminal_counts.get("generic_exit", 0)),
    }

    counts["BE after TP1 (direct)"] = int(be_after_tp_direct.get(1, 0))
    counts["BE after TP2 (direct)"] = int(be_after_tp_direct.get(2, 0))
    counts["BE after TP3 (direct)"] = int(be_after_tp_direct.get(3, 0))
    for i in range(len(tp_pcts)):
        counts[f"TP{i+1} (trades reached)"] = int(tp_reached_counts[i])

    pct_trades_initial_dd = float(len([x for x in initial_dd_values if x > 0]) / n_trades * 100.0) if n_trades else 0.0
    avg_initial_dd = float(np.mean(initial_dd_values)) if initial_dd_values else 0.0
    median_initial_dd = float(np.median(initial_dd_values)) if initial_dd_values else 0.0

    avg_fees = float(np.mean(fees_list)) if fees_list else 0.0
    tot_fees = float(np.sum(fees_list)) if fees_list else 0.0
    
    return {
        "n_trades": int(n_trades),
        "winrate": float(winrate),
        "compounded_net_profit_pct": float(compounded_net_profit_pct),
        "avg_win_pct": float(avg_win),
        "avg_loss_pct": float(avg_loss),
        "max_drawdown_pct": float(max_dd),
        "profit_factor": float(profit_factor),
        "pct_trades_initial_dd": float(pct_trades_initial_dd),
        "avg_initial_dd_pct": float(avg_initial_dd),
        "median_initial_dd_pct": float(median_initial_dd),
        "avg_fees_notional_pct": avg_fees,
        "total_fees_notional_pct": tot_fees,
        **lev_stats,
        **fixed_totals,
        "counts": counts,
    }

def _dd_ok(sl, be, L, windows, dd_limit, tables_by_sheet):

    for win in windows:

        df_window = tables_by_sheet.get(f"Last {win}D")

        if df_window is None:
            raise ValueError(f"Missing table for {win}D")

        r = _pick_row(df_window, sl, be, "NET")

        if r is None:
            return False

        d = r.get(f"Max Drawdown L{L}x (%)")

        if d is None or float(d) > dd_limit:
            return False

    return True



def evaluate_scenario(trades: List[Trade], df: pd.DataFrame,
                    tp_pcts: List[float], tp_weights: List[float],
                    sl_pct: float, sl_to_be_trigger: Optional[int],
                    initial_dd_hours: float = 2.0) -> Dict[str, Any]:
    """
    Szenario-Auswertung:
    - uses simulate_single_trade() for the core logic (incl. TP/SL/BE)
    - counts terminal closings PER TRADE (SL, BE, full TP, generic exit, partial TP)
    - counts TP1..TP6 reach per trade (baseline without BE)
    - computes compounding, MaxDD and fixed-size sums for L1x / L3x / L5x / L7x
    - computes initial drawdown metrics over the first initial_dd_hours after entry
    """

    import numpy as np
    import collections

    pnls: list[float] = []
    # Count terminal types PER TRADE
    terminal_counts: collections.Counter[str] = collections.Counter()
    # TP-Reach (pro Trade, baseline ohne BE) – TP1..TP6
    tp_reach = [0] * 6

    # BE-after-TP buckets (exklusiv, direkt: TP# erreicht und dann BE, ohne weiteres TP)
    be_after_tp_direct = {1: 0, 2: 0, 3: 0}

    # Initial-Drawdown (erste N Stunden nach Entry)
    initial_dds: list[float] = []

    # pro Trade simulieren
    for tr in trades:
        window = df[(df["ts"] >= tr.entry_time) & (df["ts"] <= tr.exit_time)][["open", "high", "low", "close", "ts"]]

        # Initial drawdown over initial_dd_hours
        if initial_dd_hours is not None and initial_dd_hours > 0:
            cutoff_ts = tr.entry_time + pd.Timedelta(hours=initial_dd_hours)
            win_init = window[window["ts"] <= cutoff_ts]
        else:
            win_init = window

        if not win_init.empty:
            if tr.side == "long":
                min_price = float(win_init["low"].min())
                dd_pct = max(0.0, (tr.entry_price - min_price) / tr.entry_price * 100.0)
            else:
                max_price = float(win_init["high"].max())
                dd_pct = max(0.0, (max_price - tr.entry_price) / tr.entry_price * 100.0)
        else:
            dd_pct = 0.0
        initial_dds.append(dd_pct)

        # Baseline-Outcome ohne BE – nur, um TP-Hits / partial events zu kennen
        base = _baseline_outcome(window, tr.side, tr.entry_price, tp_pcts, tp_weights, sl_pct)
        # Count TP reach per trade (TP1..TP6)
        for i, hit in enumerate(base["tp_hits"]):
            if i < len(tp_reach) and hit:
                tp_reach[i] += 1

        # ECHTE Simulation mit SL-to-BE etc.
        res = simulate_single_trade(window, tr.side, tr.entry_price,
                                    tp_pcts, tp_weights, sl_pct, sl_to_be_trigger)
        pnl = float(res["realized_pnl_pct"])
        pnls.append(pnl)

        br = res.get("breakdown", {})

        # Determine terminal type PER TRADE (not counting events!)
        if br.get("SL", 0) > 0:
            terminal = "SL"
        elif br.get("BE", 0) > 0:
            terminal = "BE"
        elif br.get("full_TP", 0) > 0:
            terminal = "full_TP"
        elif br.get("generic_exit", 0) > 0:
            terminal = "generic_exit"
        elif br.get("partial_TP", 0) > 0:
            terminal = "partial_TP"
        else:
            terminal = "unknown"

        # Wenn Trade am BE endet: welcher hoechste TP wurde tatsaechlich ausgefuehrt?
        if terminal == "BE":
            tp_exec = res.get("tp_exec_weights") or []
            hits = [i+1 for i,w in enumerate(tp_exec) if w is not None and float(w) > 0.0]
            if hits:
                h = max(hits)
                if h in be_after_tp_direct:
                    be_after_tp_direct[h] += 1

        terminal_counts[terminal] += 1

    pnls = np.array(pnls, dtype=float)
    n_trades = len(pnls)

    # Initial-DD Aggregation
    initial_dds = np.array(initial_dds, dtype=float) if initial_dds else np.array([], dtype=float)
    if len(initial_dds) > 0:
        pct_trades_initial_dd = float((initial_dds > 0).sum() / len(initial_dds) * 100.0)
        avg_initial_dd = float(np.mean(initial_dds))
        median_initial_dd = float(np.median(initial_dds))
    else:
        pct_trades_initial_dd = float("nan")
        avg_initial_dd = float("nan")
        median_initial_dd = float("nan")

    # Basisstatistiken zu pnls (1x)
    if n_trades == 0:
        base = {
            "n_trades": 0,
            "winrate": 0.0,
            "compounded_net_profit_pct": 0.0,
            "avg_win_pct": 0.0,
            "avg_loss_pct": 0.0,
            "max_drawdown_pct": 0.0,
            "profit_factor": 0.0,
            "pct_trades_initial_dd": pct_trades_initial_dd,
            "avg_initial_dd_pct": avg_initial_dd,
            "median_initial_dd_pct": median_initial_dd,
            "counts": {},
        }

        for L in LEVERAGES:
            base[f"L{L}x_compounded_pct"] = 0.0
            base[f"L{L}x_max_drawdown_pct"] = 0.0
            base[f"FSum_L{L}x_pct"] = 0.0

        return base


    wins = (pnls > 0).sum()
    losses = (pnls < 0).sum()
    winrate = wins / n_trades if n_trades > 0 else 0.0

    avg_win = float(pnls[pnls > 0].mean()) if wins > 0 else 0.0
    avg_loss = float(pnls[pnls < 0].mean()) if losses > 0 else 0.0

    gross_profit = pnls[pnls > 0].sum()
    gross_loss = -pnls[pnls < 0].sum()  # positiv
    profit_factor = (gross_profit / gross_loss) if gross_loss > 0 else float("inf")

    # Equity curve & MaxDD for L1x — vectorized
    stats1x = _equity_stats_from_returns(pnls)
    compounded_net_profit_pct = stats1x["compounded_pct"]
    max_dd_pct = stats1x["max_dd_pct"]

    # Leverage-Varianten (compounded & MaxDD) — vectorized
    lev_stats = {}
    for L in LEVERAGES:
        rL = np.maximum(pnls * float(L), -100.0)
        st = _equity_stats_from_returns(rL)
        lev_stats[f"L{L}x_compounded_pct"] = st["compounded_pct"]
        lev_stats[f"L{L}x_max_drawdown_pct"] = st["max_dd_pct"]

    # Fixed-Size-Summen pro Leverage
    fixed_totals = {}
    for L in LEVERAGES:
        rL = np.clip(pnls * L, -100.0, None)
        fixed_totals[f"FSum_L{L}x_pct"] = float(rL.sum())

    # Build counts dict
    counts: dict[str, int] = {}
    counts["SL"] = terminal_counts.get("SL", 0)
    counts["BE"] = terminal_counts.get("BE", 0)
    counts["partial_TP"] = terminal_counts.get("partial_TP", 0)
    counts["full_TP"] = terminal_counts.get("full_TP", 0)
    counts["generic_exit"] = terminal_counts.get("generic_exit", 0)

    # BE closes after TP1/TP2/TP3 (direct)
    counts["BE after TP1 (direct)"] = int(be_after_tp_direct.get(1, 0))
    counts["BE after TP2 (direct)"] = int(be_after_tp_direct.get(2, 0))
    counts["BE after TP3 (direct)"] = int(be_after_tp_direct.get(3, 0))

    # TP-Reach pro Trade
    for i in range(len(tp_reach)):
        counts[f"TP{i+1}_reached_trades"] = tp_reach[i]

    return {
        "n_trades": int(n_trades),
        "winrate": float(winrate),
        "compounded_net_profit_pct": float(compounded_net_profit_pct),
        "avg_win_pct": float(avg_win),
        "avg_loss_pct": float(avg_loss),
        "max_drawdown_pct": float(max_dd_pct),
        "profit_factor": float(profit_factor),
        "pct_trades_initial_dd": float(pct_trades_initial_dd),
        "avg_initial_dd_pct": float(avg_initial_dd),
        "median_initial_dd_pct": float(median_initial_dd),
        **lev_stats,
        **fixed_totals,
        "counts": counts,
    }


def _scenarios_to_table(report: dict) -> pd.DataFrame:
    rows = []

    def _add_row(sl_val, be_val, basis, res):
        counts = res.get("counts", {})

        row = {
            "Basis": basis,
            "SL (%)": sl_val,
            "BE (after TP #)": be_val,
            "Trades": res.get("n_trades", 0),
            "Winrate (%)": res.get("winrate", 0.0) * 100.0,
            "Avg Win (%)": res.get("avg_win_pct", float("nan")),
            "Avg Loss (%)": res.get("avg_loss_pct", float("nan")),
            "Profit Factor": res.get("profit_factor", float("nan")),
            "Compounded Net Profit L1x (%)": res.get("L1x_compounded_pct", float("nan")),
            "Max Drawdown L1x (%)": res.get("L1x_max_drawdown_pct", float("nan")),
            "Compounded Net Profit L2x (%)": res.get("L2x_compounded_pct", float("nan")),
            "Max Drawdown L2x (%)": res.get("L2x_max_drawdown_pct", float("nan")),
            "Compounded Net Profit L3x (%)": res.get("L3x_compounded_pct", float("nan")),
            "Max Drawdown L3x (%)": res.get("L3x_max_drawdown_pct", float("nan")),
            "Compounded Net Profit L4x (%)": res.get("L4x_compounded_pct", float("nan")),
            "Max Drawdown L4x (%)": res.get("L4x_max_drawdown_pct", float("nan")),
            "Compounded Net Profit L5x (%)": res.get("L5x_compounded_pct", float("nan")),
            "Max Drawdown L5x (%)": res.get("L5x_max_drawdown_pct", float("nan")),
            "Compounded Net Profit L6x (%)": res.get("L6x_compounded_pct", float("nan")),
            "Max Drawdown L6x (%)": res.get("L6x_max_drawdown_pct", float("nan")),
            "Compounded Net Profit L7x (%)": res.get("L7x_compounded_pct", float("nan")),
            "Max Drawdown L7x (%)": res.get("L7x_max_drawdown_pct", float("nan")),
            "Compounded Net Profit L10x (%)": res.get("L10x_compounded_pct", float("nan")),
            "Max Drawdown L10x (%)": res.get("L10x_max_drawdown_pct", float("nan")),
            "Compounded Net Profit L12x (%)": res.get("L12x_compounded_pct", float("nan")),
            "Max Drawdown L12x (%)": res.get("L12x_max_drawdown_pct", float("nan")),
            "Compounded Net Profit L15x (%)": res.get("L15x_compounded_pct", float("nan")),
            "Max Drawdown L15x (%)": res.get("L15x_max_drawdown_pct", float("nan")),
        }

        # --- Dynamic Fixed Size Columns (order preserved after compounding) ---
        for L in LEVERAGES:
            row[f"Fixed-Size Sum L{L}x (%)"] = res.get(
                f"FSum_L{L}x_pct", float("nan")
            )

        # Continue normal columns
        row["Pct trades w/ initial DD (%)"] = res.get("pct_trades_initial_dd", float("nan"))
        row["Avg initial DD (%)"] = res.get("avg_initial_dd_pct", float("nan"))

        # ---- Unified count handling (GROSS + NET compatible) ----
        row["Closed: SL"] = counts.get("Closed: SL", counts.get("SL", 0))
        row["Closed: BE"] = counts.get("Closed: BE", counts.get("BE", 0))

        row["BE after TP1 (direct)"] = counts.get("BE after TP1 (direct)", 0)
        row["BE after TP2 (direct)"] = counts.get("BE after TP2 (direct)", 0)
        row["BE after TP3 (direct)"] = counts.get("BE after TP3 (direct)", 0)

        row["Closed: partial TP"] = counts.get("Closed: partial TP", counts.get("partial_TP", 0))
        row["Closed: full TP"] = counts.get("Closed: full TP", counts.get("full_TP", 0))
        row["Closed: generic exit (pure)"] = counts.get(
            "Closed: generic exit (pure)",
            counts.get("generic_exit", 0)
        )

        # TP reached compatibility
        for i in range(1, 7):
            row[f"TP{i} (trades reached)"] = counts.get(
                f"TP{i} (trades reached)",
                counts.get(f"TP{i}_reached_trades", 0)
            )


        row["Avg Fees (notional %)"] = res.get("avg_fees_notional_pct", 0.0)
        row["Total Fees (notional %)"] = res.get("total_fees_notional_pct", 0.0)

        rows.append(row)

    for key, payload in report["scenarios"].items():
        sl_val = None
        be_val = None

        if key.startswith("SL"):
            try:
                sl_part, be_part = key.split("_")
                sl_val = float(sl_part[2:])
                be_raw = be_part[2:]
                be_val = None if be_raw == "None" else int(be_raw)
            except Exception:
                pass

        if isinstance(payload, dict) and "GROSS" in payload and "NET" in payload:
            _add_row(sl_val, be_val, "GROSS", payload["GROSS"])
            _add_row(sl_val, be_val, "NET", payload["NET"])
        else:
            _add_row(sl_val, be_val, "GROSS", payload)

    return pd.DataFrame(rows)

def _export_table(df: pd.DataFrame, csv_path: str=None, xlsx_path: str=None, html_path: str=None):
    if csv_path:
        df.to_csv(csv_path, index=False)
    if xlsx_path:
        with pd.ExcelWriter(xlsx_path, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="Scenarios")
            wb  = writer.book
            ws  = writer.sheets["Scenarios"]

            # Basic column formats
            colmap = {c:i for i,c in enumerate(df.columns)}

            def num_to_col(n: int) -> str:
                s = ""
                n += 1
                while n:
                    n, r = divmod(n-1, 26)
                    s = chr(65+r) + s
                return s

            nrows = len(df) + 1  # incl header
            def col_range(colname: str) -> str:
                idx = colmap[colname]
                col_letter = num_to_col(idx)
                return f"{col_letter}2:{col_letter}{nrows}"

            # Width + numeric/percent formats
            fmt_pct = wb.add_format({"num_format": '0.00"%"'})
            fmt_num = wb.add_format({"num_format": "0.00"})
            fmt_int = wb.add_format({"num_format": "0"})

            for col, idx in colmap.items():
                # Percent columns: header contains '%'
                if "%" in col:
                    ws.set_column(idx, idx, 16, fmt_pct)
                elif col in ["Profit Factor"]:
                    ws.set_column(idx, idx, 14, fmt_num)
                else:
                    ws.set_column(idx, idx, 14)

            # Integer-like columns
            int_cols = [
                "Trades","Closed: SL","Closed: BE","Closed: partial TP","Closed: full TP",
                "Closed: generic exit (pure)",
                "TP1 (trades reached)","TP2 (trades reached)","TP3 (trades reached)",
                "TP4 (trades reached)","TP5 (trades reached)","TP6 (trades reached)",
            ]
            for col in int_cols:
                if col in colmap:
                    ws.set_column(colmap[col], colmap[col], 14, fmt_int)

            # Shorter headers in Excel only (CSV stays unchanged)
            header_renames = {
                "Compounded Net Profit L1x (%)": "Comp. Profit L1x",
                "Compounded Net Profit L2x (%)": "Comp. Profit L2x",
                "Compounded Net Profit L3x (%)": "Comp. Profit L3x",
                "Compounded Net Profit L4x (%)": "Comp. Profit L4x",
                "Compounded Net Profit L5x (%)": "Comp. Profit L5x",
                "Compounded Net Profit L6x (%)": "Comp. Profit L6x",
                "Compounded Net Profit L7x (%)": "Comp. Profit L7x",
                "Compounded Net Profit L10x (%)": "Comp. Profit L10x",
                "Compounded Net Profit L12x (%)": "Comp. Profit L12x",
                "Compounded Net Profit L15x (%)": "Comp. Profit L15x",
                "Max Drawdown L1x (%)": "max. DD L1x",
                "Max Drawdown L2x (%)": "max. DD L2x",
                "Max Drawdown L3x (%)": "max. DD L3x",
                "Max Drawdown L4x (%)": "max. DD L4x",
                "Max Drawdown L5x (%)": "max. DD L5x",
                "Max Drawdown L6x (%)": "max. DD L6x",
                "Max Drawdown L7x (%)": "max. DD L7x",
                "Max Drawdown L10x (%)": "max. DD L10x",
                "Max Drawdown L12x (%)": "max. DD L12x",
                "Max Drawdown L15x (%)": "max. DD L15x",
                "Fixed-Size Sum L1x (%)": "Fixed-Size L1x",
                "Fixed-Size Sum L2x (%)": "Fixed-Size L2x",
                "Fixed-Size Sum L3x (%)": "Fixed-Size L3x",
                "Fixed-Size Sum L4x (%)": "Fixed-Size L4x",
                "Fixed-Size Sum L5x (%)": "Fixed-Size L5x",
                "Fixed-Size Sum L6x (%)": "Fixed-Size L6x",
                "Fixed-Size Sum L7x (%)": "Fixed-Size L7x",
                "Fixed-Size Sum L10x (%)": "Fixed-Size L10x",
                "Fixed-Size Sum L12x (%)": "Fixed-Size L12x",
                "Fixed-Size Sum L15x (%)": "Fixed-Size L15x",
            }
            for col, idx in colmap.items():
                if col in header_renames:
                    ws.write(0, idx, header_renames[col])

            # Conditional formatting: high is good
            good_high = [
                "Winrate (%)",
                "Compounded Net Profit L1x (%)",
                "Compounded Net Profit L2x (%)",
                "Compounded Net Profit L3x (%)",
                "Compounded Net Profit L4x (%)",
                "Compounded Net Profit L5x (%)",
                "Compounded Net Profit L6x (%)",
                "Compounded Net Profit L7x (%)",
                "Compounded Net Profit L10x (%)",
                "Compounded Net Profit L12x (%)",
                "Compounded Net Profit L15x (%)",
                "Fixed-Size Sum L1x (%)",
                "Fixed-Size Sum L2x (%)",
                "Fixed-Size Sum L3x (%)",
                "Fixed-Size Sum L4x (%)",
                "Fixed-Size Sum L5x (%)",
                "Fixed-Size Sum L6x (%)",
                "Fixed-Size Sum L7x (%)",
                "Fixed-Size Sum L10x (%)",
                "Fixed-Size Sum L12x (%)",
                "Fixed-Size Sum L15x (%)",
                "Profit Factor",
            ]
            for col in good_high:
                if col in colmap:
                    ws.conditional_format(col_range(col), {
                        "type": "3_color_scale",
                        "min_color": "#f8696b",  # rot
                        "mid_color": "#ffeb84",  # gelb
                        "max_color": "#63be7b",  # green
                    })

            # Conditional formatting: low is good
            good_low = [
                "Max Drawdown L1x (%)",
                "Max Drawdown L2x (%)",
                "Max Drawdown L3x (%)",
                "Max Drawdown L4x (%)",
                "Max Drawdown L5x (%)",
                "Max Drawdown L6x (%)",
                "Max Drawdown L7x (%)",
                "Max Drawdown L10x (%)",
                "Max Drawdown L12x (%)",
                "Max Drawdown L15x (%)",
                "Trades w/ initial DD (%)",
                "Avg initial DD (%)",
            ]
            for col in good_low:
                if col in colmap:
                    ws.conditional_format(col_range(col), {
                        "type": "3_color_scale",
                        "min_color": "#63be7b",  # green
                        "mid_color": "#ffeb84",  # gelb
                        "max_color": "#f8696b",  # rot
                    })

            # Autofilter & freeze
            ws.autofilter(0, 0, len(df), len(df.columns)-1)
            ws.freeze_panes(1, 0)
    if html_path:
        styled = (df.style
                    .format({
                        "Winrate (%)": "{:.2f}%",
                        "Avg Win (%)": "{:.2f}%",
                        "Avg Loss (%)": "{:.2f}%",
                        "Profit Factor": "{:.2f}",
                    })
                    .set_table_styles([
                        {"selector":"th","props":[("background","#f4f6f8"),("text-align","center"),("padding","6px")]},
                        {"selector":"td","props":[("padding","6px")]},
                        {"selector":"table","props":[("border-collapse","collapse"),("border","1px solid #ddd")]},
                    ])
                    .hide(axis="index"))
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(styled.to_html())

from openpyxl.styles import Border, Side

thick = Side(style="medium")

def border_row(ws, row, left_col, right_col, *, top=False, bottom=False):
    for c in range(left_col, right_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = Border(
            top=thick if top else None,
            bottom=thick if bottom else None,
        )

def border_lr(ws, row, left_col, right_col):
    ws.cell(row=row, column=left_col).border = Border(left=thick)
    ws.cell(row=row, column=right_col).border = Border(right=thick)



def _equity_stats_from_returns(pnls_pct) -> dict:
    """Given list/array of per-trade returns in %, compute compounded net profit % and max drawdown %.
    Vectorized with NumPy cumprod — ~10x faster than the original loop.
    """
    arr = np.asarray(pnls_pct, dtype=float) / 100.0
    if arr.size == 0:
        return {"equity": 1.0, "compounded_pct": 0.0, "max_dd_pct": 0.0}
    curve = np.cumprod(1.0 + arr)
    equity = float(curve[-1])
    peak = np.maximum.accumulate(curve)
    dd = (peak - curve) / peak
    max_dd = float(dd.max())
    return {
        "equity": equity,
        "compounded_pct": (equity - 1.0) * 100.0,
        "max_dd_pct": max_dd * 100.0
    }

# -------------------------------
# CLI
# -------------------------------


def _compute_report(trades: List[Trade],
                    df: pd.DataFrame,
                    tp: List[float],
                    w: List[float],
                    sl_list: List[float],
                    be_list: List[Any],
                    initial_dd_hours: float,
                    maker_fee_pct: float,
                    taker_fee_pct: float,):
    """
    Compute scenario results for a given trade list, keeping the core logic identical
    to the main run. Returns a report dict with the same structure as before.
    """
    # --- HARD STRUCTURE VALIDATION ---
    if len(tp) != 6:
        raise ValueError(f"System requires exactly 6 TP levels. Got {len(tp)}")

    if len(w) != 6:
        raise ValueError(f"System requires exactly 6 TP weights. Got {len(w)}")

    if abs(sum(w) - 1.0) > 1e-6:
        raise ValueError(
            f"TP weights must sum to 1.0. Current sum = {sum(w):.6f}"
        )

    report: Dict[str, Any] = {"tp": tp, "w": w, "scenarios": {}}

    for sl in sl_list:
        for be in be_list:
            key = f"SL{sl}_BE{be}"
            res_gross = evaluate_scenario(trades, df, tp, w, sl, be, initial_dd_hours=initial_dd_hours)
            res_net = evaluate_scenario_net(trades, df, tp, w, sl, be, initial_dd_hours=initial_dd_hours,
                                        maker_fee_pct=maker_fee_pct, taker_fee_pct=taker_fee_pct,)
            report["scenarios"][key] = {"GROSS": res_gross, "NET": res_net}
    return report


def _export_table_excel_multi(dfs_by_sheet: Dict[str, pd.DataFrame], xlsx_path: str, params: Optional[Dict[str, Any]] = None):
    """
    Export multiple scenario tables into one Excel file, each DataFrame into its own sheet,
    using the same formatting/conditional formatting as the single-sheet export.
    """
    if not xlsx_path:
        return

    with pd.ExcelWriter(xlsx_path, engine="xlsxwriter") as writer:
        wb = writer.book

        # Formats (same as single-sheet)
        fmt_pct = wb.add_format({"num_format": '0.00"%"'})
        fmt_num = wb.add_format({"num_format": "0.00"})
        fmt_int = wb.add_format({"num_format": "0"})

        # Color scales (same logic as before)
        # Green (high) -> Red (low)
        fmt_scale_good_high = {"type": "3_color_scale"}
        # Red (high) -> Green (low)
        fmt_scale_good_low = {"type": "3_color_scale", "min_color": "#63BE7B", "mid_color": "#FFEB84", "max_color": "#F8696B"}

        for sheet_name, df in dfs_by_sheet.items():
            if df is None:
                continue
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]

            # Header renames (short names) – same mapping as in _export_table
            header_renames = {
                "Compounded Net Profit L1x (%)": "Comp. Profit L1x",
                "Compounded Net Profit L2x (%)": "Comp. Profit L2x",
                "Compounded Net Profit L3x (%)": "Comp. Profit L3x",
                "Compounded Net Profit L4x (%)": "Comp. Profit L4x",
                "Compounded Net Profit L5x (%)": "Comp. Profit L5x",
                "Compounded Net Profit L6x (%)": "Comp. Profit L6x",
                "Compounded Net Profit L7x (%)": "Comp. Profit L7x",
                "Compounded Net Profit L10x (%)": "Comp. Profit L10x",
                "Compounded Net Profit L12x (%)": "Comp. Profit L12x",
                "Compounded Net Profit L15x (%)": "Comp. Profit L15x",
                "Max Drawdown L1x (%)": "max. DD L1x",
                "Max Drawdown L2x (%)": "max. DD L2x",
                "Max Drawdown L3x (%)": "max. DD L3x",
                "Max Drawdown L4x (%)": "max. DD L4x",
                "Max Drawdown L5x (%)": "max. DD L5x",
                "Max Drawdown L6x (%)": "max. DD L6x",
                "Max Drawdown L7x (%)": "max. DD L7x",
                "Max Drawdown L10x (%)": "max. DD L10x",
                "Max Drawdown L12x (%)": "max. DD L12x",
                "Max Drawdown L15x (%)": "max. DD L15x",
                "Fixed-Size Sum L1x (%)": "Fixed-Size L1x",
                "Fixed-Size Sum L2x (%)": "Fixed-Size L2x",
                "Fixed-Size Sum L3x (%)": "Fixed-Size L3x",
                "Fixed-Size Sum L4x (%)": "Fixed-Size L4x",
                "Fixed-Size Sum L5x (%)": "Fixed-Size L5x",
                "Fixed-Size Sum L6x (%)": "Fixed-Size L6x",
                "Fixed-Size Sum L7x (%)": "Fixed-Size L7x",
                "Fixed-Size Sum L10x (%)": "Fixed-Size L10x",
                "Fixed-Size Sum L12x (%)": "Fixed-Size L12x",
                "Fixed-Size Sum L15x (%)": "Fixed-Size L15x",
            }

            # Apply header renames in first row
            for col_idx, col_name in enumerate(df.columns):
                display = header_renames.get(col_name, col_name)
                ws.write(0, col_idx, display)

            # Build column map
            colmap = {col: i for i, col in enumerate(df.columns)}
            nrows = len(df) + 1  # + header row

            def num_to_col(n: int) -> str:
                s = ""
                n0 = n
                while True:
                    n0, r = divmod(n0, 26)
                    s = chr(ord("A") + r) + s
                    if n0 == 0:
                        break
                    n0 -= 1
                return s

            def col_range(colname: str) -> str:
                idx = colmap[colname]
                col_letter = num_to_col(idx)
                return f"{col_letter}2:{col_letter}{nrows}"

            # Width + numeric/percent formats
            for col, idx in colmap.items():
                if "%" in col:
                    ws.set_column(idx, idx, 16, fmt_pct)
                elif col in ["n_trades", "Closed_SL", "Closed_BE", "Closed_partial_TP", "Closed_full_TP",
                            "Closed_generic_exit_pure", "TP1_reached_trades", "TP2_reached_trades", "TP3_reached_trades",
                            "TP4_reached_trades", "TP5_reached_trades", "TP6_reached_trades"]:
                    ws.set_column(idx, idx, 14, fmt_int)
                else:
                    ws.set_column(idx, idx, 16, fmt_num)

            # Filters & freeze
            ws.autofilter(0, 0, nrows - 1, len(df.columns) - 1)
            ws.freeze_panes(1, 0)

            # Conditional formatting (only if columns exist)
            good_high = [
                "Winrate (%)",
                "Compounded Net Profit L1x (%)",
                "Compounded Net Profit L2x (%)",
                "Compounded Net Profit L3x (%)",
                "Compounded Net Profit L4x (%)",
                "Compounded Net Profit L5x (%)",
                "Compounded Net Profit L6x (%)",
                "Compounded Net Profit L7x (%)",
                "Compounded Net Profit L10x (%)",
                "Compounded Net Profit L12x (%)",
                "Compounded Net Profit L15x (%)",
                "Fixed-Size Sum L1x (%)",
                "Fixed-Size Sum L2x (%)",
                "Fixed-Size Sum L3x (%)",
                "Fixed-Size Sum L4x (%)",
                "Fixed-Size Sum L5x (%)",
                "Fixed-Size Sum L6x (%)",
                "Fixed-Size Sum L7x (%)",
                "Fixed-Size Sum L10x (%)",
                "Fixed-Size Sum L12x (%)",
                "Fixed-Size Sum L15x (%)",
                "Profit Factor",
            ]
            good_low = [
                "Max Drawdown L1x (%)",
                "Max Drawdown L2x (%)",
                "Max Drawdown L3x (%)",
                "Max Drawdown L4x (%)",
                "Max Drawdown L5x (%)",
                "Max Drawdown L6x (%)",
                "Max Drawdown L7x (%)",
                "Max Drawdown L10x (%)",
                "Max Drawdown L12x (%)",
                "Max Drawdown L15x (%)",
                "Trades w/ initial DD (%)",
                "Avg initial DD (%)",
            ]

            for colname in good_high:
                if colname in colmap:
                    ws.conditional_format(col_range(colname), fmt_scale_good_high)

            for colname in good_low:
                if colname in colmap:
                    ws.conditional_format(col_range(colname), fmt_scale_good_low)
        # --- Parameters sheet (TP ladder, fees, input file) ---
        if params:
            sheet_params_name = params.get("sheet_name", "Parameters")
            # avoid duplicate
            if sheet_params_name in writer.sheets:
                ws_p = writer.sheets[sheet_params_name]
            else:
                ws_p = wb.add_worksheet(sheet_params_name)
                writer.sheets[sheet_params_name] = ws_p

            bold = wb.add_format({"bold": True})
            fmt_txt = wb.add_format({})
            fmt_pct = wb.add_format({"num_format": '0.00"%"'})
            fmt_dt = wb.add_format({"num_format": "yyyy-mm-dd hh:mm"})

            # Header
            ws_p.write(0, 0, "Backtest Parameters Summary", bold)

            # Meta
            row = 2

            csv_name = Path(params.get("input_csv", "")).name

            ws_p.write(row, 0, "Input CSV", bold)
            ws_p.write(row, 1, csv_name, fmt_txt)
            row += 1

            ws_p.write(row, 0, "Baseline", bold)
            ws_p.write(row, 1, str(params.get("input_baseline", "")), fmt_txt)
            row += 1

            ws_p.write(row, 0, "Sensitivity", bold)
            ws_p.write(row, 1, str(params.get("input_sensitivity", "")), fmt_txt)
            row += 1

            ws_p.write(row, 0, "Timeframe", bold)
            ws_p.write(row, 1, str(params.get("input_timeframe", "")), fmt_txt)
            row += 1

            ws_p.write(row, 0, "Stoploss Type", bold)
            ws_p.write(row, 1, str(params.get("stoploss_type", "")), fmt_txt)
            row += 1

            ws_p.write(row, 0, "Stoploss Type", bold)
            ws_p.write(row, 1, str(params.get("stoploss_type", "")), fmt_txt)
            row += 1

            ws_p.write(row, 0, "Stoploss Value", bold)
            ws_p.write(row, 1, params.get("stoploss_value", ""), fmt_txt)
            row += 1

            ws_p.write(row, 0, "Stoploss Value", bold)
            ws_p.write(row, 1, params.get("stoploss_value", ""), fmt_txt)
            row += 1

            generated_dt = params.get("generated_utc")

            if isinstance(generated_dt, str):
                try:
                    generated_dt = datetime.fromisoformat(generated_dt)
                except Exception:
                    generated_dt = None

            if generated_dt:
                generated_dt = generated_dt.replace(tzinfo=None)
                ws_p.write_datetime(row, 1, generated_dt, fmt_dt)
            else:
                ws_p.write(row, 1, "", fmt_txt)

            ws_p.write(row, 0, "Generated (UTC)", bold)
            row += 2



            # Fees/block
            ws_p.write(row, 0, "Maker fee (%)", bold); ws_p.write_number(row, 1, float(params.get("maker_fee_pct", 0.0)), fmt_pct); row += 1
            ws_p.write(row, 0, "Taker fee (%)", bold); ws_p.write_number(row, 1, float(params.get("taker_fee_pct", 0.0)), fmt_pct); row += 1

            # TP ladder table
            ws_p.write(row, 0, "TP Ladder", bold); row += 1
            ws_p.write(row, 0, "TP #", bold)
            ws_p.write(row, 1, "Target (%)", bold)
            ws_p.write(row, 2, "Weight", bold)
            row += 1

            tp = params.get("tp", []) or []
            wts = params.get("w", []) or []
            for i in range(min(len(tp), len(wts))):
                ws_p.write_number(row, 0, i + 1, fmt_int)
                ws_p.write_number(row, 1, float(tp[i]), fmt_pct)
                ws_p.write_number(row, 2, float(wts[i]), wb.add_format({"num_format": "0.000"}))
                row += 1

            # Column widths
            ws_p.set_column(0, 0, 10)
            ws_p.set_column(1, 1, 18)
            ws_p.set_column(2, 2, 14)
            ws_p.set_column(0, 2, 22)


def _compute_final_profile_suggestions(tables_by_sheet):
    """
    Returns dict:
    {
        "SAFE": {...},
        "BALANCED": {...},
        "AGGRESSIVE": {...}
    }
    """

    df30   = tables_by_sheet.get("Last 30D")
    df90   = tables_by_sheet.get("Last 90D")
    df180  = tables_by_sheet.get("Last 180D")
    df360  = tables_by_sheet.get("Last 360D")

    # Require core windows
    if df30 is None or df90 is None or df180 is None:
        return None

# 360D is optional but preferred


    profiles = {
        "SAFE": 20.0,
        "BALANCED": 30.0,
        "AGGRESSIVE": 40.0,
    }

    results = {}

    for profile_name, dd_limit in profiles.items():

        best_score = -1
        best_combo = None

        # anchor on largest window
        anchor_window = max(WINDOWS)
        df_anchor = tables_by_sheet.get(f"Last {anchor_window}D")

        if df_anchor is None:
            continue


        for _, r_anchor in df_anchor.iterrows():

            if r_anchor["Basis"] != "NET":
                continue

            sl = r_anchor["SL (%)"]
            be = r_anchor["BE (after TP #)"]

            rows = {}

            valid = True
            for win in WINDOWS:
                dfw = tables_by_sheet.get(f"Last {win}D")
                if dfw is None:
                    valid = False
                    break

                match = dfw[
                    (dfw["SL (%)"] == sl) &
                    (dfw["BE (after TP #)"] == be) &
                    (dfw["Basis"] == "NET")
                ]

                if len(match) == 0:
                    valid = False
                    break

                rows[win] = match.iloc[0]

            if not valid:
                continue

            score = _profile_master_score(rows, dd_limit)

            if score and score > best_score:
                best_score = score
                best_combo = {
                    "sl": sl,
                    "be": be,
                    "lev": r_anchor["Leverage"],
                    "score": score
                }

        if best_combo:
            results[profile_name] = best_combo

    return results

# ============================================================
# MASTER PROFILE SCORE (30/90/180/360 GEOMETRIC GROWTH UNDER DD)
# ============================================================

def _profile_master_score(rows_dict, dd_limit):
    """
    rows_dict: { window_int: row }
    """

    growth_factors = []

    for win, row in rows_dict.items():

        if row is None:
            return None

        try:
            cnp = float(row["Compounded Net Profit L1x (%)"])
            dd  = float(row["Max Drawdown L1x (%)"])

        except Exception:
            return None

        # Strict DD filter
        if dd > dd_limit:
            return None

        g = 1 + cnp / 100.0

        if g <= 0:
            return None

        growth_factors.append(g)

    if not growth_factors:
        return None

    # Geometric mean across all windows
    product = 1.0
    for g in growth_factors:
        product *= g

    return product ** (1.0 / len(growth_factors))



def _add_top5_profitfactor_sheet(
    xlsx_path: str,
    tables_by_sheet: dict,
    trades_by_sheet: dict,
    df: pd.DataFrame,
    tp: list[float],
    w: list[float],
    initial_dd_hours: float,
    maker_fee_pct: float,
    taker_fee_pct: float,
    rank_window: str = "180D",
    baseline: str = "",
    sensitivity: str = "",
    timeframe: str = "",
    input_csv: str = "",
) -> None:

    """
    Append a 'Top5 ProfitFactor (180D)' sheet to an existing scenario xlsx.

    NEW:
    - Top6 (ranked by Profit Factor on 180D NET, BE != None)
    - Layout: 3 blocks per row (TOP1-3) + (TOP4-6) below, like the provided template
    - Avg Win/Loss + Wins/Losses in a Row are recomputed per window & basis (GROSS/NET)
    - Global conditional formatting across all TOP6:
        Winrate / ProfitFactor / CNP -> Green-to-Red
        MDD -> Red-to-Green
    - DD-threshold fallback: still show values; if fallback is used (DD > threshold), font is red
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.formatting.rule import ColorScaleRule
    except Exception as e:
        print(f"[WARN] openpyxl not available for Top5 sheet: {e}")
        return

    if not xlsx_path:
        return

    if rank_window not in ("30D", "90D", "180D", "360D"):
        raise ValueError("rank_window must be '30D', '90D', '180D' or '360D'")


    token, tf = _infer_token_and_tf_from_csv(input_csv)
    token_name, token_tf = _infer_token_and_tf_from_csv(input_csv)


    def _pick_row(dfX, sl, be, basis):
        if dfX is None:
            return None
        m = (dfX["SL (%)"] == sl) & (dfX["BE (after TP #)"] == be) & (dfX["Basis"] == basis)
        sub = dfX.loc[m]
        if len(sub) == 0:
            return None
        return sub.iloc[0]



    def _read_cell(ws, addr, default=""):
        try:
            v = ws[addr].value
            return "" if v is None else str(v)
        except Exception:
            return default

    # Leverage sets and DD thresholds
    risk_cfg = {
        "Safe": {
            "dd": 20.0,
            "levs": [1, 2, 3, 4],
            "min_lev": 1,
        },
        "Balanced": {
            "dd": 30.0,
            "levs": [3, 4, 5, 6, 7],
            "min_lev": 2,
        },
        "Aggressive": {
            "dd": 40.0,
            "levs": [5, 6, 7, 10, 12, 15],
            "min_lev": 3,
        },
    }


    def _has_top1_strict_chain(sl, be, windows, risk_cfg, tables_by_sheet):
        """
        Returns True if there exists a STRICT Safe→Balanced→Aggressive
        leverage chain satisfying DD across all required windows.
        """

        for Ls in risk_cfg["Safe"]["levs"]:
            if not _dd_ok(sl, be, Ls, windows, risk_cfg["Safe"]["dd"], tables_by_sheet):
                continue

            for Lb in risk_cfg["Balanced"]["levs"]:
                if Lb < Ls:
                    continue
                if not _dd_ok(sl, be, Lb, windows, risk_cfg["Balanced"]["dd"], tables_by_sheet):
                    continue

                for La in risk_cfg["Aggressive"]["levs"]:
                    if La < Lb:
                        continue
                    if not _dd_ok(sl, be, La, windows, risk_cfg["Aggressive"]["dd"], tables_by_sheet):
                        continue

                    return True

        return False

    def _has_strict_monotonic_leverage(sl, be, windows, rank_window, strict: bool, tables_by_sheet):

        prev_lev = 0

        for label in ("Safe", "Balanced", "Aggressive"):

            cfg = risk_cfg[label]
            allowed = [L for L in cfg["levs"] if L > prev_lev]

            found = False

            for L in sorted(allowed, reverse=True):

                ok = True

                for win in windows:

                    df_window = tables_by_sheet.get(f"Last {win}D")

                    if df_window is None:
                        raise ValueError(f"Missing table for {win}D")

                    r = _pick_row(df_window, sl, be, "NET")

                    if r is None:
                        ok = False
                        break

                    d = r.get(f"Max Drawdown L{L}x (%)")

                    if d is None or float(d) > cfg["dd"]:
                        ok = False
                        break

                if ok:
                    prev_lev = L
                    found = True
                    break

            if not found:
                return False

        return True


    rank_map = {
        f"{w}D": tables_by_sheet.get(f"Last {w}D")
        for w in WINDOWS
    }

    df_rank_src = rank_map.get(rank_window)
    if df_rank_src is None:
        print(f"[WARN] Top5 sheet skipped: missing Last {rank_window} table")
        return

    # Rank candidates by Profit Factor (rank_window NET) and BE != None
    df_rank = df_rank_src.copy()
    df_rank = df_rank[
        (df_rank["Basis"] == "NET") &
        (df_rank["BE (after TP #)"].notna())
    ]
    df_rank = df_rank.sort_values(by="Profit Factor", ascending=False)

    top_rows = []
    seen = set()

    if rank_window == "30D":
        validity_windows = (30,)
    elif rank_window == "90D":
        validity_windows = (30, 90)
    elif rank_window == "180D":
        validity_windows = (30, 90, 180)
    elif rank_window == "360D":
        validity_windows = (90, 180, 360)   # ← 30D removed here
    else:
        validity_windows = tuple(WINDOWS)



    for _, r in df_rank.iterrows():
        try:
            sl = float(r["SL (%)"])
            be = int(r["BE (after TP #)"])
        except Exception:
            continue

        key = (sl, be)
        if key in seen:
            continue

        # STRICT admission: full monotonic chain must exist
        if not _has_top1_strict_chain(sl, be, validity_windows, risk_cfg, tables_by_sheet):
            continue

        top_rows.append(key)
        seen.add(key)

        if len(top_rows) == 6:
            break

    if not top_rows:
        raise RuntimeError(
            "[STRICT MODE] No candidates satisfy full Safe→Balanced→Aggressive "
            "DD constraints."
        )

        return

    # ============================================================
    # FINAL PROFILE MASTER SUGGESTIONS (30/90/180/360)
    # ============================================================

    def _compute_profile_suggestion(dd_limit):

        best_score = -1
        best_combo = None

        # Use largest window as anchor (latest history)
        anchor_window = max(WINDOWS)
        df_anchor = tables_by_sheet.get(f"Last {anchor_window}D")

        if df_anchor is None:
            return None

        for _, r_anchor in df_anchor.iterrows():

            if r_anchor["Basis"] != "NET":
                continue

            try:
                sl = float(r_anchor["SL (%)"])
                be = int(r_anchor["BE (after TP #)"])
            except Exception:
                continue

            # Collect rows across all windows
            rows = {}

            for win in WINDOWS:
                df_window = tables_by_sheet.get(f"Last {win}D")
                if df_window is None:
                    return None

                rows[win] = _pick_row(df_window, sl, be, "NET")

            score = _profile_master_score(rows, dd_limit)

            if score and score > best_score:
                best_score = score
                best_combo = {
                    "sl": sl,
                    "be": be,
                    "score": score
                }

        return best_combo



    profile_suggestions = {
        "Safe": _compute_profile_suggestion(20.0),
        "Balanced": _compute_profile_suggestion(30.0),
        "Aggressive": _compute_profile_suggestion(40.0),
    }

    # --------------------------------
    # PASS 2: STRICT TOP2–TOP6
    # --------------------------------
    for _, r in df_rank.iterrows():
        if len(top_rows) >= 6:
            break

        try:
            sl = float(r["SL (%)"])
            be = int(r["BE (after TP #)"])
        except Exception:
            continue

        key = (sl, be)
        if key in seen:
            continue

        if not _has_top1_strict_chain(sl, be, validity_windows, risk_cfg, tables_by_sheet):
            continue

        top_rows.append(key)
        seen.add(key)
    
    
    def _compute_streaks_and_avgs(pnls: list[float]):
        """
        pnls are per-trade realized pnl pct (already gross or net, in percent terms).
        Returns: (wins_in_a_row, losses_in_a_row, avg_win, avg_loss)
        avg_win and avg_loss in % (loss is negative number, like existing table convention).
        """
        best_w = 0
        best_l = 0
        cur_w = 0
        cur_l = 0

        wins = []
        losses = []

        for x in pnls:
            try:
                v = float(x)
            except Exception:
                continue

            if v > 0:
                wins.append(v)
                cur_w += 1
                cur_l = 0
            elif v < 0:
                losses.append(v)
                cur_l += 1
                cur_w = 0
            else:
                # treat 0 as break in both
                cur_w = 0
                cur_l = 0

            best_w = max(best_w, cur_w)
            best_l = max(best_l, cur_l)

        avg_win = sum(wins) / len(wins) if wins else 0.0
        avg_loss = sum(losses) / len(losses) if losses else 0.0
        return best_w, best_l, avg_win, avg_loss

    def _simulate_trade_pnls_for_basis(trades, sl, be, basis: str):
        """
        Re-simulate trade-by-trade using simulate_single_trade to compute:
        - gross pnl
        - net pnl (+fees) consistent with backtester fee logic

        Returns list of pnl% per trade.
        """
        pnls_out = []
        if not trades:
            return pnls_out

        for tr in trades:
            try:
                window = df[(df["ts"] >= tr.entry_time) & (df["ts"] <= tr.exit_time)][["open", "high", "low", "close", "ts"]]
            except Exception:
                continue

            res = simulate_single_trade(
                window, tr.side, tr.entry_price,
                tp, w, sl, be
            )

            gross_pnl = float(res.get("realized_pnl_pct", 0.0))
            if basis == "GROSS":
                pnls_out.append(gross_pnl)
                continue

            # NET
            # We need these fields to do maker/taker by exit-type:
            # - tp_exec_weights: list of weights actually executed at TP levels
            # - remainder_exit_weight or remainder_exit_weight (fallback names)
            # - close_reason
            tp_exec = res.get("tp_exec_weights", None)
            if tp_exec is None:
                # backwards compat: if missing, approximate with entry taker + exit taker on full notional
                entry_fee = float(taker_fee_pct)
                exit_fee = float(taker_fee_pct)
                fees_notional = entry_fee + exit_fee
                dur_hours = max(0.0, (tr.exit_time - tr.entry_time).total_seconds() / 3600.0)
                continue

            rem = float(res.get("remainder_exit_weight", 0.0))
            rem = float(rem) if rem is not None else 0.0
            close_reason = str(res.get("close_reason", "generic_exit"))

            maker_notional = float(sum(float(x) for x in tp_exec)) if tp_exec else 0.0

            # entry fee: taker on full notional
            entry_fee = float(taker_fee_pct)

            # remainder exit fee by terminal type
            if close_reason == "BE":
                remainder_fee_rate = float(maker_fee_pct)     # maker-like for BE close
            elif close_reason in ("SL", "generic_exit"):
                remainder_fee_rate = float(taker_fee_pct)     # taker-like
            else:
                remainder_fee_rate = 0.0

            exit_fee = maker_notional * float(maker_fee_pct) + rem * remainder_fee_rate
            fees_notional = entry_fee + exit_fee


            net_pnl = gross_pnl - fees_notional
            pnls_out.append(net_pnl)

        return pnls_out

    # workbook/sheet
    wb = load_workbook(xlsx_path)
    sheet_name = f"Top6 ProfitFactor ({rank_window})"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)

    title_font = Font(bold=True, size=14)
    hdr_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    fill_hdr = PatternFill("solid", fgColor="D9E1F2")
    red_font = Font(color="9C0006")  # dark red
    danger_font = Font(color="9C0006", bold=True, underline="single")

    headers = ["Metric"]
    col_pairs = []

    for win in WINDOWS:
        headers.append(f"{win} GROSS")
        headers.append(f"{win} NET")
        col_pairs.append((win, "GROSS"))
        col_pairs.append((win, "NET"))


    # Template-like layout: 3 blocks per row
    # Blocks start at C, J, Q (1-based: 3, 10, 17)
    DATA_COLS_PER_WINDOW = 2
    BLOCK_WIDTH = 1 + len(WINDOWS) * DATA_COLS_PER_WINDOW
    GAP = 1  # 1 column between blocks

    block_start_cols = [
        3 + i * (BLOCK_WIDTH + GAP)
        for i in range(6)
    ]


    top_section_row = 3
    bottom_section_row = None  # computed later

    # All metrics (order matches your template)
    metric_order = [
        "SL (%)",
        "BE (after TP #)",
        "Winrate (%)",
        "Profit Factor",
        "Safe-CNP",
        "Safe-MDD",
        "Safe-LEV",
        "Balanced-CNP",
        "Balanced-MDD",
        "Balanced-LEV",
        "Aggressive-CNP",
        "Aggressive-MDD",
        "Aggressive-LEV",
        "Wins in a Row",
        "Losses in a Row",
        "Avg Win (%)",
        "Avg Loss (%)",
        "BE after TP1 (direct)",
        "BE after TP2 (direct)",
        "BE after TP3 (direct)",
        "Trades with generic Exit",
        "Trades with SL",
        "Trades reached TP1",
        "Trades reached TP2",
        "Trades reached TP3",
        "Trades reached TP4",
        "Trades reached TP5",
        "Trades reached TP6",
    ]
    print("\n[RANK CHECK]")
    for i, (sl, be) in enumerate(top_rows, 1):
        r = _pick_row(df_rank_src, sl, be, "NET")
        if r is None:
            print(f"TOP{i}: SL={sl} BE={be} → MISSING")
        else:
            print(f"TOP{i}: SL={sl} BE={be} → PF={r['Profit Factor']}")

    # helper: fill a block at (sr, sc) for TOPk
    def _write_block(rank_idx, sl, be, sr, sc):
        # collect window table rows (for risk best-of + counts)
        rows = {}

        DATA_COLS_PER_WINDOW = 2  # GROSS + NET
        BLOCK_WIDTH = 1 + len(WINDOWS) * DATA_COLS_PER_WINDOW

        for win in WINDOWS:
            dfx = tables_by_sheet.get(f"Last {win}D")
            if dfx is None:
                continue

            for basis in ("GROSS", "NET"):
                rows[(win, basis)] = _pick_row(dfx, sl, be, basis)


        # windows that must respect DD for this sheet
        if rank_window == "30D":
            validity_windows = (30,)
        elif rank_window == "90D":
            validity_windows = (30, 90)
        elif rank_window == "180D":
            validity_windows = (30, 90, 180)
        elif rank_window == "360D":
            validity_windows = (90, 180, 360)   # ⬅ 30D excluded
        else:
            validity_windows = tuple(WINDOWS)

        rows_net_by_win = {}
        for win in validity_windows:
            rows_net_by_win[win] = rows.get((win, "NET"))

        # trade lists per window
        win_trades_map = {
            win: trades_by_sheet.get(f"Last {win}D", [])
            for win in WINDOWS
        }

        for win in WINDOWS:
            label = f"Trades ({win}D)"

        # recompute streaks + avg win/loss per (win,basis)
        streak_cache = {}
        for win in (WINDOWS):
            tlist = win_trades_map.get(win, [])
            for basis in ("GROSS", "NET"):
                pnls = _simulate_trade_pnls_for_basis(tlist, sl, be, basis=basis)
                wins_row, losses_row, avg_win, avg_loss = _compute_streaks_and_avgs(pnls)
                streak_cache[(win, basis)] = {
                    "wins_row": wins_row,
                    "losses_row": losses_row,
                    "avg_win": avg_win,
                    "avg_loss": avg_loss,
                }

        def gv(win, basis, col, default=float("nan")):
            r = rows.get((win, basis))
            if r is None:
                return default
            return r.get(col, default)

        # title
        from openpyxl.styles import Font, Alignment, PatternFill

        title_row = sr

        # --- TOP badge (bigger, centered, background) ---
        top_cell = ws.cell(row=title_row, column=sc, value=f"TOP {rank_idx}")
        ws.merge_cells(
            start_row=title_row,
            start_column=sc,
            end_row=title_row,
            end_column=sc + 1
        )

        top_cell.font = Font(bold=True, size=16)
        top_cell.alignment = Alignment(horizontal="center", vertical="center")
        top_cell.fill = PatternFill("solid", fgColor="c4daff")  # light blue
        size_map = {
            1: 18,
            2: 16,
            3: 16,
            4: 15,
            5: 15,
            6: 15,
        }

        color_map = {
            1: "000000",
            2: "222222",
            3: "222222",
            4: "444444",
            5: "444444",
            6: "444444",
        }

        top_cell.font = Font(
            name="Calibri",
            size=size_map.get(rank_idx, 15),
            bold=True,
            underline="single",
            color=color_map.get(rank_idx, "444444")
        )

        # --- SL ---
        sl_cell = ws.cell(
            row=title_row,
            column=sc + 2,
            value=f"SL={sl}%"
        )
        sl_cell.font = Font(bold=True, size=12, color="555555")
        sl_cell.alignment = Alignment(horizontal="left", vertical="center")

        # --- BE ---
        be_cell = ws.cell(
            row=title_row,
            column=sc + 4,
            value=f"BE after TP={be}"
        )
        be_cell.font = Font(bold=True, size=12, color="555555")
        be_cell.alignment = Alignment(horizontal="left", vertical="center")

        from openpyxl.styles import Border, Side

        thick = Side(style="medium")

        above_title_row = sr - 1
        left_col = sc
        right_col = sc + BLOCK_WIDTH - 1  # same width you use for the title/meta box

        for c in range(left_col, right_col + 1):
            ws.cell(row=above_title_row, column=c).border = Border(bottom=thick)

        # ---- Meta strip (right of title, per-block safe) ----
        meta_row = sr + 1

        meta_items = [
            ("Baseline", baseline),
            ("Signal sens", sensitivity),
            ("TF", timeframe),
        ]

        for i, (label, value) in enumerate(meta_items):
            col = sc + i * 2   # 👈 uses sc, sc+2, sc+4 (inside block)
            cell = ws.cell(
                row=meta_row,
                column=col,
                value=f"{label}: {value}"
            )
            cell.font = Font(bold=True, size=12, color="555555")
            cell.alignment = center

        from openpyxl.styles import Border, Side

        thick = Side(style="medium")   # use "thin" if you prefer
        thin  = Side(style="thin")

        top_row = sr
        bottom_row = sr + 1
        left_col = sc
        block_width_actual = len(headers)
        right_col = sc + BLOCK_WIDTH - 1


        # left border
        for r in (top_row, bottom_row):
            ws.cell(row=r, column=left_col).border = Border(left=thick)

        # right border
        for r in (top_row, bottom_row):
            ws.cell(row=r, column=right_col).border = Border(right=thick)
        fill = PatternFill("solid", fgColor="c4daff")

        for r in (top_row, bottom_row):
            for c in range(left_col, right_col + 1):
                ws.cell(row=r, column=c).fill = fill

        from openpyxl.styles import Border, Side

        thick = Side(style="medium")


        # header row
        hr = sr + 2
        for j, h in enumerate(headers):
            c = ws.cell(row=hr, column=sc + j, value=h)
            c.font = hdr_font
            c.fill = fill_hdr
            c.alignment = center
        # ---- top border on header row (closes the title/meta box) ----
        for c in range(sc, sc + len(headers)):
            ws.cell(row=hr, column=c).border = Border(top=thick)


        # metric rows start
        mr0 = sr + 4
        trades_row = mr0 - 1

        # ---- Trades row (window-only, not basis-specific) ----
        col_offset = 1
        for win in WINDOWS:
            trades_list = win_trades_map.get(win, [])
            ws.cell(row=trades_row, column=sc + col_offset, value=len(trades_list))
            ws.cell(row=trades_row, column=sc + col_offset).alignment = center
            ws.cell(row=trades_row, column=sc + col_offset).number_format = "0"
            col_offset += 2


        # leave NET columns empty on purpose
        col_offset = 2
        for _ in WINDOWS:
            ws.cell(row=trades_row, column=sc + col_offset, value="")
            col_offset += 2

        def _pick_leverage_strict(
            rows_net_by_win,
            levs,
            min_lev,
            dd_limit,
            prev_lev,
            validity_windows,
        ):
            """
            Strict leverage picker.

            Returns:
                lev or None

            Rules:
            - Enforces monotonicity (>= prev_lev)
            - Enforces minimum leverage
            - MUST satisfy DD limit across ALL validity_windows
            - No fallback, no extended thresholds
            """

            allowed = [L for L in levs if L >= max(min_lev, prev_lev)]

            def dd_ok(L):
                for win in validity_windows:
                    r = rows_net_by_win.get(win)
                    if r is None:
                        return False
                    d = r.get(f"Max Drawdown L{L}x (%)")
                    if d is None or float(d) > dd_limit:
                        return False
                return True

            for L in sorted(allowed, reverse=True):
                if dd_ok(L):
                    return L

            return None



        # -----------------------------
        # Risk selection & aggregation (STRICT)
        # -----------------------------

        risk_vals = {}
        risk_lev  = {}

        # Pre-fill risk_vals
        for lbl in ("Safe", "Balanced", "Aggressive"):
            for win in (WINDOWS):
                for basis in ("GROSS", "NET"):
                    risk_vals[(lbl, "CNP", win, basis)] = float("nan")
                    risk_vals[(lbl, "MDD", win, basis)] = float("nan")

        prev_lev = 0  # enforce monotonic Safe ≤ Balanced ≤ Aggressive
        if rank_window == "30D":
            validity_windows = (30,)
        elif rank_window == "90D":
            validity_windows = (30, 90)
        elif rank_window == "180D":
            validity_windows = (30, 90, 180)
        elif rank_window == "360D":
            validity_windows = (90, 180, 360)
        else:
            validity_windows = tuple(WINDOWS)


        for lbl in ("Safe", "Balanced", "Aggressive"):

            cfg = risk_cfg[lbl]

            rows_net_by_win = {
                win: rows.get((win, "NET"))
                for win in validity_windows
            }

            lev = _pick_leverage_strict(
                rows_net_by_win=rows_net_by_win,
                levs=cfg["levs"],
                min_lev=cfg["min_lev"],
                dd_limit=cfg["dd"],
                prev_lev=prev_lev,
                validity_windows=validity_windows,
            )

            if lev is None:
                raise RuntimeError(
                    f"[SAFETY] {lbl} cannot satisfy DD limits "
                    f"(prev={prev_lev})"
                )

            risk_lev[lbl] = lev
            prev_lev = lev

            # Apply same leverage across all windows & bases
            for win in (WINDOWS):
                for basis in ("GROSS", "NET"):
                    r = rows.get((win, basis))
                    if r is None:
                        continue
                    risk_vals[(lbl, "CNP", win, basis)] = r.get(
                        f"Compounded Net Profit L{lev}x (%)"
                    )
                    risk_vals[(lbl, "MDD", win, basis)] = r.get(
                        f"Max Drawdown L{lev}x (%)"
                    )


        # ---- 3️⃣ compute per-window DD violations (visual-only) ----
        dd_violation = {}  # (label, win) -> bool

        for lbl in ("Safe", "Balanced", "Aggressive"):
            lev = risk_lev.get(lbl)
            cfg = risk_cfg[lbl]

            for win in (WINDOWS):
                r = rows.get((win, "NET"))
                if r is None or lev is None:
                    dd_violation[(lbl, win)] = False
                    continue

                d = r.get(f"Max Drawdown L{lev}x (%)")
                dd_violation[(lbl, win)] = (
                    d is not None and float(d) > cfg["dd"]
                )

            
        # write each metric row
        for i, mname in enumerate(metric_order):
            rrow = mr0 + i
            ws.cell(row=rrow, column=sc, value=mname).alignment = left

            for j, (win, basis) in enumerate(col_pairs, start=1):
                cell = ws.cell(row=rrow, column=sc + j)

                # values
                if mname == "SL (%)":
                    val = float(sl)
                elif mname == "BE (after TP #)":
                    val = int(be)
                elif mname == "Winrate (%)":
                    val = gv(win, basis, "Winrate (%)")
                elif mname == "Profit Factor":
                    val = gv(win, basis, "Profit Factor")   
                elif mname == "Avg Win (%)":
                    val = float(streak_cache[(win, basis)]["avg_win"])
                elif mname == "Avg Loss (%)":
                    val = float(streak_cache[(win, basis)]["avg_loss"])
                elif mname == "Wins in a Row":
                    val = int(streak_cache[(win, basis)]["wins_row"])
                elif mname == "Losses in a Row":
                    val = int(streak_cache[(win, basis)]["losses_row"])
                elif mname.startswith("BE after TP") and "(direct)" in mname:
                    k = int(mname.split("TP")[1].split()[0])
                    val = gv(win, basis, f"BE after TP{k} (direct)", 0)
                elif mname == "Trades with generic Exit":
                    val = gv(win, basis, "Closed: generic exit (pure)", 0)
                elif mname == "Trades with SL":
                    val = gv(win, basis, "Closed: SL", 0)
                elif mname.startswith("Trades reached TP"):
                    k = int(mname.split("TP")[1])
                    val = gv(win, basis, f"TP{k} (trades reached)", 0)
                elif mname.endswith("-CNP"):
                    label = mname.split("-")[0]
                    val = risk_vals[(label, "CNP", win, basis)]                  

                elif mname.endswith("-MDD"):
                    label = mname.split("-")[0]
                    val = risk_vals[(label, "MDD", win, basis)]                  

                elif mname.endswith("-LEV"):
                    label = mname.split("-")[0]
                    val = risk_lev.get(label)
                    
                
                else:
                    val = float("nan")

                cell.value = val
                

        # formatting for columns within this block
        # Percent formatting: SL, Winrate, Avg Win/Loss, all CNP/MDD
        pct_metrics = {
            "SL (%)", "Winrate (%)", "Avg Win (%)", "Avg Loss (%)",
            "Safe-CNP", "Safe-MDD",
            "Balanced-CNP", "Balanced-MDD",
            "Aggressive-CNP", "Aggressive-MDD",
        }
        for i, mname in enumerate(metric_order):
            rr = mr0 + i
            for j in range(1, len(col_pairs) + 1):  # data columns
                c = ws.cell(row=rr, column=sc + j)
                if mname in pct_metrics:
                    c.number_format = '0.00"%"'
                elif mname == "Profit Factor":
                    c.number_format = '0.00'
                elif mname in ("SL (%)",):
                    c.number_format = '0.00"%"'
                elif mname.endswith("-LEV"):
                    c.number_format = "0"
                    
                else:
                    # integers for counts/streaks
                    if mname in (
                        "Wins in a Row", "Losses in a Row",
                        "Trades with generic Exit", "Trades with SL",
                        "BE after TP1 (direct)", "BE after TP2 (direct)", "BE after TP3 (direct)",
                        "Trades reached TP1", "Trades reached TP2", "Trades reached TP3",
                        "Trades reached TP4", "Trades reached TP5", "Trades reached TP6",
                    ):
                        c.number_format = '0'

        # return useful row indices for CF mapping
        return {
            "title_row": sr,
            "header_row": hr,
            "metric_start": mr0,
            "metric_rows": {name: (mr0 + idx) for idx, name in enumerate(metric_order)},
            "block_left": sc,
            "block_right": sc + 6,
        }

    # sheet title
    ws.cell(row=1, column=3, value=f"Top6 variants (ranked by Profit Factor on {rank_window} NET, BE != None)").font = title_font
    ws.cell(row=1, column=10, value=f"Token: {token_name}").font = Font(bold=True, size=16, color="444444")

    # UX note: how many DD-valid strategies were found
    ws.cell(row=2, column=3, value=f"Note: {len(top_rows)} DD-valid strategies found for {rank_window}. "
        f"TOP1 is strictly within base limits; TOP2–TOP6 may use extended DD (shown in red)." ).font = Font(italic=True, color="666666")

    # write top 3 blocks
    blocks_meta = []
    top_count = len(top_rows)

    # write top section (up to 3)
    for i in range(min(3, top_count)):
        sl, be = top_rows[i]
        meta = _write_block(i + 1, sl, be, top_section_row, block_start_cols[i])
        blocks_meta.append(meta)


    # determine bottom section row (block height + spacer)
    # block height = 1(title)+1(header)+len(metric_order)
    block_height = 3 + len(metric_order)
    bottom_section_row = top_section_row + block_height + 3  # spacer rows between sections

    # write bottom 3 blocks (if available)
    if top_count > 3:
        for i in range(3, min(6, top_count)):
            sl, be = top_rows[i]
            col_idx = i - 3
            meta = _write_block(i + 1, sl, be, bottom_section_row, block_start_cols[col_idx])
            blocks_meta.append(meta)


    # column widths for the entire sheet (roughly like template)
    for c in range(1, 35):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 18
    # wider "Metric" columns inside blocks
    for sc in block_start_cols:
        ws.column_dimensions[ws.cell(row=1, column=sc).column_letter].width = 24

    # Global conditional formatting across all TOP6 blocks
    # We'll apply to full span C:V for the specific metric row in each section.
    # Because spacer columns are blank, Excel ignores them for min/max.
    def _add_colorscale(range_str: str, reverse: bool = False):
        if not reverse:
            rule = ColorScaleRule(
                start_type="min", start_color="F8696B",   # red
                mid_type="percentile", mid_value=50, mid_color="FFEB84",  # yellow
                end_type="max", end_color="63BE7B"       # green
            )
        else:
            rule = ColorScaleRule(
                start_type="min", start_color="63BE7B",  # green
                mid_type="percentile", mid_value=50, mid_color="FFEB84", # yellow
                end_type="max", end_color="F8696B"       # red
            )
        ws.conditional_formatting.add(range_str, rule)

    # Identify metric rows (top and bottom) by using the first block meta (layout identical)
    top_rows_map = blocks_meta[0]["metric_rows"]

    bottom_rows_map = (
        blocks_meta[3]["metric_rows"]
        if len(blocks_meta) > 3
        else None
    )

    from xlsxwriter.utility import xl_rowcol_to_cell

    def _ranges_for_metric(metric_name: str, top: bool):

        if top:
            r = top_rows_map[metric_name]
        else:
            if bottom_rows_map is None:
                return []
            r = bottom_rows_map[metric_name]

        DATA_COLS_PER_WINDOW = 2
        BLOCK_WIDTH = 1 + len(WINDOWS) * DATA_COLS_PER_WINDOW
        GAP = 1

        # Excel rows are 1-based, xlsxwriter is 0-based
        row_idx = r - 1

        ranges = []

        for i in range(3):  # 3 blocks per row (TOP1–3 or TOP4–6)

            start_col = 2 + i * (BLOCK_WIDTH + GAP)  # column C = 2
            end_col = start_col + BLOCK_WIDTH - 1

            start_cell = xl_rowcol_to_cell(row_idx, start_col)
            end_cell = xl_rowcol_to_cell(row_idx, end_col)

            ranges.append(f"{start_cell}:{end_cell}")

        return ranges

    # rows with green-to-red (higher is better)
    gr_metrics = ["Winrate (%)", "Profit Factor", "Safe-CNP", "Balanced-CNP", "Aggressive-CNP"]
    # rows with red-to-green (lower is better)
    rg_metrics = ["Safe-MDD", "Balanced-MDD", "Aggressive-MDD"]

    for mname in gr_metrics:
        for rng in _ranges_for_metric(mname, top=True):
            _add_colorscale(rng, reverse=False)
        for rng in _ranges_for_metric(mname, top=False):
            _add_colorscale(rng, reverse=False)

    for mname in rg_metrics:
        for rng in _ranges_for_metric(mname, top=True):
            _add_colorscale(rng, reverse=True)
        for rng in _ranges_for_metric(mname, top=False):
            _add_colorscale(rng, reverse=True)


    wb.save(xlsx_path)
    print(f"[OK] Added sheet '{sheet_name}' to: {xlsx_path}")
    return True


def main():
    p = argparse.ArgumentParser(description="TP/SL Ladder Simulator")
    p.add_argument("--csv", required=True, help="CSV with OHLCV + Signals (time, open, high, low, close, Buy Entry Signal, Sell Entry Signal, Exit Signal)")
    p.add_argument("--selection_window", default="180")
    p.add_argument("--input_baseline", default="", help="Metadata only: baseline label (written to Parameters sheet)")
    p.add_argument("--input_sensitivity", default="", help="Metadata only: sensitivity value (written to Parameters sheet)")
    p.add_argument("--input_timeframe", default="", help="Metadata only: timeframe label (written to Parameters sheet)")
    p.add_argument("--tp", required=False, default="[2,4,6]", help="TP ladder in %%, e.g. [2,4,6]")
    p.add_argument("--w", required=False, default="[0.5,0.3,0.2]", help="Weights for TPs (sum ~ 1.0), e.g. [0.5,0.3,0.2]")
    p.add_argument("--sl_list", required=False, default="[2,3,5]", help="List of SL levels in %%, e.g. [2,3,5]")
    p.add_argument("--be_list", required=False, default="[null,1,2,3]", help="List of SL-to-BE triggers: null or 1/2/3...")
    p.add_argument("--initial_dd_hours", required=False, type=float, default=2.0,
                help="Time period in hours after entry over which the initial drawdown is measured (e.g. 2.0).")
    p.add_argument("--maker_fee_pct", type=float, default=0.015,
                help="Maker fee in %% of notional per fill (default 0.015).")
    p.add_argument("--taker_fee_pct", type=float, default=0.025,
                help="Taker fee in %% of notional per fill (default 0.025).")
    p.add_argument("--out", required=False, default="", help="Optional: path for JSON report")
    p.add_argument("--table_csv", required=False, default="", help="Optional: CSV of comparison table")
    p.add_argument("--table_xlsx", required=False, default="", help="Optional: Excel table (formatted)")
    p.add_argument("--table_html", required=False, default="", help="Optional: HTML table (formatted)")
    p.add_argument("--stoploss_type", default="")
    p.add_argument("--stoploss_value", type=float, default=None)

    args = p.parse_args()

    # Build dynamic window list based on selection_window —
    # only include windows that are <= the requested lookback.
    # Always keep 30D as the shortest context window.
    _sel_win = int(args.selection_window)
    _all_windows = [30, 90, 180, 360]
    WINDOWS[:] = [w for w in _all_windows if w <= _sel_win]
    # Always include at least 30D even if selection_window < 30
    if not WINDOWS:
        WINDOWS[:] = [30]
    print(f"📅 Active windows: {[str(w)+'D' for w in WINDOWS]} (selection_window={_sel_win}D)")

    tp = json.loads(args.tp)
    w  = json.loads(args.w)
    sl_list = json.loads(args.sl_list)
    be_list = json.loads(args.be_list)
    # Validate / normalize TP weights
    if len(w) != len(tp):
        print(f"[WARN] TP weights length ({len(w)}) != TP levels length ({len(tp)}). This may lead to unexpected results.")
    s_w = sum(float(x) for x in w)
    if abs(s_w - 1.0) > 1e-6:
        if s_w == 0:
            print("[WARN] TP weights sum to 0. Cannot normalize.")
        else:
            print(f"[WARN] TP weights sum is {s_w:.6f} (expected 1.0). Normalizing to 1.0.")
            w = [float(x) / s_w for x in w]

    df = load_signals_csv(args.csv)

    df = df.copy()
    df["ts"] = pd.to_datetime(df["ts"], utc=True, errors="coerce")
    df = df.dropna(subset=["ts"])
    df = df.sort_values("ts").reset_index(drop=True)

    trades = build_trades(df)

    print(f"✅ {len(trades)} trades built from signals.")

    report = _compute_report(
        trades,
        df,
        tp,
        w,
        sl_list,
        be_list,
        args.initial_dd_hours,
        args.maker_fee_pct,
        args.taker_fee_pct,
    )

    # Nicely print
    import pprint
    pp = pprint.PrettyPrinter(indent=2, width=120, sort_dicts=False)
    pp.pprint(report)

    # Build comparison table
    table_df = _scenarios_to_table(report)

    # Exports
    if args.table_csv:
        _export_table(table_df, csv_path=args.table_csv)
        print(f"\n💾 CSV table saved to: {args.table_csv}")

    # HTML bleibt wie bisher (All Trades)
    if args.table_html:
        _export_table(table_df, html_path=args.table_html)
        print(f"💾 HTML table saved to: {args.table_html}")

    # Excel: All Trades + additional sheets for Last 30/90/180 Days (Entry-Time >= Cutoff and Exit in window)
    if args.table_xlsx:
        last_ts = pd.to_datetime(df["ts"].iloc[-1], utc=True, errors="coerce")
        if pd.isna(last_ts):
            # Fallback: wenn ts nicht parsebar ist, nutze jetzt
            last_ts = pd.Timestamp.utcnow()

        def _filter_trades_window(days: int) -> List[Trade]:
            cutoff = last_ts - pd.Timedelta(days=int(days))
            out: List[Trade] = []
            for t in trades:
                try:
                    et = pd.to_datetime(t.entry_time, utc=True)
                    xt = pd.to_datetime(t.exit_time, utc=True)
                except Exception:
                    continue
                if et >= cutoff and xt >= cutoff:
                    out.append(t)
            return out

        tables_by_sheet: Dict[str, pd.DataFrame] = {"All Trades": table_df}
        trades_by_sheet: Dict[str, List[Trade]] = {"All Trades": trades}

        for days in WINDOWS:
            label = f"Last {days}D"
            tw = _filter_trades_window(days)
            trades_by_sheet[label] = tw
            rep_w = _compute_report(
            tw,
            df,
            tp,
            w,
            sl_list,
            be_list,
            args.initial_dd_hours,
            args.maker_fee_pct,
            args.taker_fee_pct,
        )
            df_w = _scenarios_to_table(rep_w)
            tables_by_sheet[label] = df_w
            print(f"🧩 Window {label}: {len(tw)} Trades (cutoff: {(last_ts - pd.Timedelta(days=int(days))).strftime('%Y-%m-%d %H:%M:%S UTC')})")

        # Parameters summary sheet (for Excel)
        from datetime import datetime
        params_summary = {
            "input_csv": args.csv,
            "generated_utc": datetime.now(timezone.utc),
            "tp": tp,
            "w": w,
            "maker_fee_pct": args.maker_fee_pct,
            "input_baseline": args.input_baseline,
            "stoploss_type": args.stoploss_type,
            "stoploss_value": args.stoploss_value,
            "input_sensitivity": args.input_sensitivity,
            "input_timeframe": args.input_timeframe,
            "taker_fee_pct": args.taker_fee_pct,
            "sheet_name": "Parameters",
        }


        _export_table_excel_multi(tables_by_sheet, args.table_xlsx, params=params_summary)
        params = params_summary  # already built earlier

        baseline     = params.get("input_baseline", "")
        sensitivity  = params.get("input_sensitivity", "")
        timeframe    = params.get("input_timeframe", "")

        # ------------------------------------------------------------
        # Add Top6 ProfitFactor summary sheets
        # ------------------------------------------------------------

        # Build rank windows dynamically — same logic as WINDOWS,
        # only include windows <= selection_window
        rank_windows = [f"{w}D" for w in WINDOWS]

        for rank_window in rank_windows:
            try:
                ok = _add_top5_profitfactor_sheet(
                    args.table_xlsx,
                    tables_by_sheet,
                    trades_by_sheet,
                    df,
                    tp,
                    w,
                    args.initial_dd_hours,
                    args.maker_fee_pct,
                    args.taker_fee_pct,
                    rank_window=rank_window,
                    baseline=baseline,
                    sensitivity=sensitivity,
                    timeframe=timeframe,
                    input_csv=args.csv,
                )

                if ok:
                    print(f"✓ Added Top6 ProfitFactor ({rank_window}) sheet")
                else:
                    print(f"⚠ Skipped Top6 ProfitFactor ({rank_window}) sheet")

            except Exception as e:
                print(f"[WARN] Could not add Top6 ProfitFactor ({rank_window}) sheet: {e}")

        print(f"💾 Excel table saved to: {args.table_xlsx}")



    if args.out:
        with open(args.out, "w") as f:
            json.dump(report, f, indent=2)
        print(f"\n💾 Report saved to: {args.out}")

if __name__ == "__main__":
    main()
