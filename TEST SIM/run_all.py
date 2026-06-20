#!/usr/bin/env python3

# EXPECTED:
# strategy_engine.py must output:
#   strategy_profiles.json with keys:
#   profiles -> safe/balanced/aggressive
#
# backtest_simulator.py must produce:
#   scenario_table_leverage_<profile>.xlsx
#   with sheet:
#   "Top6 ProfitFactor ({window}D)"

import argparse
import subprocess
import sys
import json
import shutil
import re
from pathlib import Path
from datetime import datetime, timezone

import pandas as pd
import matplotlib
matplotlib.use("Agg")   # non-interactive backend — no display needed
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================
# Utilities
# ============================================================

def run_subprocess(cmd, cwd):
    print(f"\n>> Running: {' '.join(str(c) for c in cmd)}")
    p = subprocess.run(cmd, cwd=str(cwd))
    if p.returncode != 0:
        print("❌ Subprocess failed.")
        sys.exit(p.returncode)


def round_list(x, d=4):
    return [round(float(v), d) for v in x]


def profile_changed(old, new):
    """Binary exact-match check — kept for backward compat but not used for publishing."""
    if old is None:
        return True
    for p in ["safe", "balanced", "aggressive"]:
        if round_list(old["profiles"][p]["tp"]) != round_list(new["profiles"][p]["tp"]):
            return True
        if round_list(old["profiles"][p]["w"]) != round_list(new["profiles"][p]["w"]):
            return True
        if float(old["profiles"][p]["sl"]) != float(new["profiles"][p]["sl"]):
            return True
        if old["profiles"][p]["be"] != new["profiles"][p]["be"]:
            return True
        if old["profiles"][p]["lev"] != new["profiles"][p]["lev"]:
            return True
    return False


# ============================================================
# Stability Gate
# ============================================================

# How much a value must change before we consider it meaningful.
# TP levels: relative change threshold (0.10 = 10% move in the level itself)
# Weights:   absolute change per slot (0.04 = 4 percentage points)
# SL / BE / LEV: any change triggers update (these are discrete, not continuous)

TP_CHANGE_THRESHOLD  = 0.10   # 10% relative move in any TP level
W_CHANGE_THRESHOLD   = 0.04   # 4pp absolute move in any weight slot


def stability_gate(old, new) -> tuple:
    """
    Compare old and new strategy profiles using meaningful thresholds
    rather than exact equality.

    Returns:
        (should_update: bool, reason: str)

    Rules:
      - If no active strategy exists          → always update
      - If SL, BE, or LEV changed             → update (discrete values)
      - If any TP level moved > 10% relative  → update
      - If any weight slot moved > 4pp        → update
      - Otherwise                             → hold, no update

    The 10% TP threshold means TP1 at 2.00% would need to move
    outside 1.80%–2.20% before triggering. A neutral week where
    MFE shifts by 0.1% would NOT trigger.

    The 4pp weight threshold means a weight at 20% would need to
    move outside 16%–24% before triggering.
    """
    if old is None:
        return True, "No active strategy — first run"

    reasons = []

    for p in ["safe", "balanced", "aggressive"]:

        # Skip profile if missing from either strategy
        if p not in old.get("profiles", {}):
            reasons.append(f"{p}: not in previous strategy")
            continue
        if p not in new.get("profiles", {}):
            reasons.append(f"{p}: not in new strategy")
            continue

        old_p = old["profiles"][p]
        new_p = new["profiles"][p]

        # Discrete values — any change is meaningful
        if float(old_p["sl"]) != float(new_p["sl"]):
            reasons.append(f"{p}: SL changed {old_p['sl']} → {new_p['sl']}")

        if old_p["be"] != new_p["be"]:
            reasons.append(f"{p}: BE changed {old_p['be']} → {new_p['be']}")

        if old_p["lev"] != new_p["lev"]:
            reasons.append(f"{p}: LEV changed {old_p['lev']} → {new_p['lev']}")

        # TP levels — relative change
        old_tp = old_p["tp"]
        new_tp = new_p["tp"]
        for i, (o, n) in enumerate(zip(old_tp, new_tp)):
            o, n = float(o), float(n)
            if o > 1e-9:
                rel = abs(n - o) / o
                if rel > TP_CHANGE_THRESHOLD:
                    reasons.append(
                        f"{p}: TP{i+1} moved {rel*100:.1f}% "
                        f"({o:.2f}% → {n:.2f}%)"
                    )

        # Weights — absolute change
        old_w = old_p["w"]
        new_w = new_p["w"]
        for i, (o, n) in enumerate(zip(old_w, new_w)):
            o, n = float(o), float(n)
            if abs(n - o) > W_CHANGE_THRESHOLD:
                reasons.append(
                    f"{p}: W{i+1} moved {abs(n-o)*100:.1f}pp "
                    f"({o*100:.0f}% → {n*100:.0f}%)"
                )

    if reasons:
        return True, " | ".join(reasons)
    else:
        return False, "All changes within stability thresholds — holding active"

def normalize_weights_exact(weights, decimals=4):
    total = sum(weights)
    if total == 0:
        return weights

    # Normalize
    normalized = [w / total for w in weights]

    # Round all except last
    rounded = [round(w, decimals) for w in normalized[:-1]]

    # Last weight absorbs rounding remainder
    remainder = round(1.0 - sum(rounded), decimals)
    rounded.append(remainder)

    return rounded

def extract_top6_winner(xlsx_path, window, profile_name):
    """
    Read the Top6 ProfitFactor sheet and return the best SL/BE/LEV combo.

    Selection logic (applied to the top 3 blocks only):
      1. Read TOP1, TOP2, TOP3 — each gives SL, BE, LEV, Winrate, PF
      2. Build a "close group": TOP1 always included; TOP2/TOP3 included
         if their Profit Factor is within 5% of TOP1's PF
         (i.e. PF >= TOP1_PF * 0.95)
      3. Within the close group sort by:
           a) Highest Winrate  (selection window NET)
           b) Lowest SL        (tiebreaker)
      4. Return the winner of that sort as the active strategy combo.

    If only TOP1 is readable, it is returned as-is (no tiebreaker needed).
    Falls back to the first valid block if PF data is unavailable.
    """
    sheet     = f"Top6 ProfitFactor ({window}D)"
    label_col = 2
    lev_label = f"{profile_name.capitalize()}-LEV"

    # Column indices for the first 3 TOP blocks (0-based pandas columns)
    # Blocks are at Excel cols 3, 10, 17 → pandas 0-based: 2, 9, 16
    # data value column is one to the right of the label column inside each block
    # label_col=2 (Excel col C), value starts at Excel col D=3 → pandas col 3
    candidate_value_cols = [3, 10, 17, 24, 31, 38]   # all 6 blocks

    df = pd.read_excel(
        xlsx_path,
        sheet_name=sheet,
        header=None,
        engine="openpyxl"
    )

    def find_value(label, value_col):
        row = df[df[label_col].astype(str).str.strip() == label]
        if row.empty:
            return None
        val = row.iloc[0, value_col] if value_col < df.shape[1] else None
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return None
        return val

    # ------------------------------------------------------------------
    # Read all candidates (up to top 3 for tiebreaker, rest as fallback)
    # ------------------------------------------------------------------
    candidates = []

    for vcol in candidate_value_cols:
        sl_val      = find_value("SL (%)",          vcol)
        be_val      = find_value("BE (after TP #)", vcol)
        lev_val     = find_value(lev_label,          vcol)
        pf_val      = find_value("Profit Factor",    vcol)
        winrate_val = find_value("Winrate (%)",      vcol)

        if sl_val is None or lev_val is None:
            continue

        try:
            candidates.append({
                "sl":      float(sl_val),
                "be":      None if pd.isna(be_val) else int(be_val),
                "lev":     int(lev_val),
                "pf":      float(pf_val)      if pf_val      is not None else 0.0,
                "winrate": float(winrate_val) if winrate_val is not None else 0.0,
            })
        except (ValueError, TypeError):
            continue

    if not candidates:
        raise ValueError(
            f"Could not extract valid SL/BE/LEV from any TOP block "
            f"in sheet '{sheet}' for profile '{profile_name}'"
        )

    # ------------------------------------------------------------------
    # Tiebreaker: check top 3 only
    # ------------------------------------------------------------------
    top3 = candidates[:3]

    if len(top3) == 1:
        # Only one candidate — use it directly
        winner = top3[0]
    else:
        top1_pf   = top3[0]["pf"]
        threshold = top1_pf * 0.95   # 5% below TOP1 PF

        # Build close group — TOP1 always included
        close_group = [c for c in top3 if c["pf"] >= threshold]

        if len(close_group) == 1:
            # No other candidate is close — TOP1 wins outright
            winner = close_group[0]
            print(f"  [Top6] {profile_name}: TOP1 wins outright "
                  f"(PF={top1_pf:.2f}, next not within 5%)")
        else:
            # Sort close group: lowest SL first (tightest risk),
            # then best winrate as tiebreaker within same SL.
            # Rationale: if PFs are within 5% of each other, always
            # prefer the tighter SL — same reward, less risk.
            close_group.sort(key=lambda c: (c["sl"], -c["winrate"]))
            winner = close_group[0]
            print(
                f"  [Top6] {profile_name}: tiebreaker applied "
                f"({len(close_group)} candidates within 5% PF of {top1_pf:.2f}) "
                f"→ selected SL={winner['sl']} BE={winner['be']} "
                f"WR={winner['winrate']:.1f}%"
            )

    return {"sl": winner["sl"], "be": winner["be"], "lev": winner["lev"]}


# ============================================================
# Main
# ============================================================


# ============================================================
# Workbook Merger — combines per-profile xlsx into one file
# ============================================================

# Sheet name prefix per profile (used as label in combined workbook)
PROFILE_LABELS = {
    "safe":       "Safe",
    "balanced":   "Balanced",
    "aggressive": "Aggressive",
}

# Tab colours per profile (green / blue / red)
PROFILE_TAB_COLORS = {
    "safe":       "22C55E",
    "balanced":   "3B82F6",
    "aggressive": "EF4444",
}

def merge_profile_workbooks(
    base_run_dir: Path,
    profiles: list,
    safe_name: str,
    selection_window: str,
    combined_profiles: dict = None,
    equity_pngs: dict = None,
) -> Path:
    """
    Read the per-profile scenario_table_{profile}.xlsx files and copy every
    sheet into a single combined workbook with prefixed sheet names and
    full colour formatting re-applied from scratch.

    Formatting scheme:
      - Header row : dark slate background (#1E293B), white bold text
      - Data rows  : alternating white / light profile tint
      - Good-high columns (Winrate, PF, Compounded profit, Fixed-size):
            green → yellow → red conditional scale
      - Good-low columns (Max Drawdown):
            red → yellow → green conditional scale
      - Top6 sheets: title/badge fills preserved (light blue #C4DAFF)
      - Tab colours: green (Safe) / blue (Balanced) / red (Aggressive)

    The individual per-profile files are left intact — extract_top6_winner()
    still reads them directly.

    Returns the path to the combined workbook.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter, column_index_from_string
    from copy import copy

    # ------------------------------------------------------------------
    # Colour palette (easy to tune — all in one place)
    # ------------------------------------------------------------------
    HDR_BG       = "1E293B"   # dark slate header background
    HDR_FG       = "FFFFFF"   # white header text

    ROW_TINTS = {             # light tint for alternating data rows
        "safe":       "F0FDF4",   # green tint
        "balanced":   "EFF6FF",   # blue tint
        "aggressive": "FFF1F2",   # red tint
    }
    ROW_WHITE    = "FFFFFF"

    # Winner row highlight — darker shade of each profile colour
    WINNER_BG = {
        "safe":       "166534",   # deep green
        "balanced":   "1E3A8A",   # deep blue
        "aggressive": "7F1D1D",   # deep red
    }
    WINNER_FG    = "FFFFFF"   # white text on dark background
    WINNER_SIZE  = 13         # font size for winner row

    # Conditional-scale colours (green→yellow→red and reverse)
    CF_GREEN     = "63BE7B"
    CF_YELLOW    = "FFEB84"
    CF_RED       = "F8696B"

    # Columns whose higher value is better (green = high)
    GOOD_HIGH_KEYWORDS = [
        "winrate", "profit factor", "compounded", "fixed-size", "comp. profit",
    ]
    # Columns whose lower value is better (green = low)
    GOOD_LOW_KEYWORDS = [
        "max drawdown", "max. dd", "drawdown",
    ]

    def _is_scenario_sheet(name: str) -> bool:
        """True for data sheets (All Trades / Last NND / Scenarios)."""
        n = name.lower()
        return ("all trades" in n or "last " in n or n == "scenarios")

    def _is_top6_sheet(name: str) -> bool:
        return "top6" in name.lower() or "top5" in name.lower()

    def _is_parameters_sheet(name: str) -> bool:
        return "parameter" in name.lower()

    def _col_is_good_high(header: str) -> bool:
        h = header.lower()
        return any(k in h for k in GOOD_HIGH_KEYWORDS)

    def _col_is_good_low(header: str) -> bool:
        h = header.lower()
        return any(k in h for k in GOOD_LOW_KEYWORDS)

    def _make_hdr_fill():
        return PatternFill("solid", fgColor=HDR_BG)

    def _make_tint_fill(profile):
        return PatternFill("solid", fgColor=ROW_TINTS[profile])

    def _make_white_fill():
        return PatternFill("solid", fgColor=ROW_WHITE)

    def _make_hdr_font():
        return Font(bold=True, color=HDR_FG, name="Calibri", size=10)

    def _make_data_font():
        return Font(name="Calibri", size=10)

    def _center():
        return Alignment(horizontal="center", vertical="center", wrap_text=False)

    def _left():
        return Alignment(horizontal="left", vertical="center")

    # ------------------------------------------------------------------
    # Build workbook
    # ------------------------------------------------------------------
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    combined_path = base_run_dir / f"{safe_name}_combined_{timestamp}.xlsx"
    combined_wb   = Workbook()
    combined_wb.remove(combined_wb.active)   # remove default blank sheet

    for profile in profiles:
        label     = PROFILE_LABELS[profile]
        tab_color = PROFILE_TAB_COLORS[profile]
        tint_fill = _make_tint_fill(profile)
        src_path  = base_run_dir / profile / f"scenario_table_{profile}.xlsx"

        if not src_path.exists():
            print(f"[WARN] Merge: {src_path.name} not found — skipping {profile}")
            continue

        try:
            # data_only=False so xlsxwriter-generated files are read correctly
            src_wb = load_workbook(src_path, data_only=False)
        except Exception as e:
            print(f"[WARN] Merge: could not open {src_path.name}: {e}")
            continue

        for src_sheet_name in src_wb.sheetnames:

            # ── Build destination sheet name ──────────────────────────
            display      = src_sheet_name.replace("Top6 ProfitFactor", "Top6")
            combined_name = f"{label} - {display}"
            if len(combined_name) > 31:
                combined_name = combined_name[:31]

            suffix    = 1
            base_name = combined_name
            while combined_name in combined_wb.sheetnames:
                combined_name = f"{base_name[:28]}_{suffix}"
                suffix += 1

            src_ws  = src_wb[src_sheet_name]
            dest_ws = combined_wb.create_sheet(title=combined_name)
            dest_ws.sheet_properties.tabColor = tab_color

            # ── Copy freeze panes ─────────────────────────────────────
            if src_ws.freeze_panes:
                dest_ws.freeze_panes = src_ws.freeze_panes

            # ── Copy column widths — auto-fit based on cell content ───
            # First copy any explicitly set widths from source
            for col_letter, col_dim in src_ws.column_dimensions.items():
                if col_dim.width and col_dim.width > 0:
                    dest_ws.column_dimensions[col_letter].width = col_dim.width

            # Then auto-fit all columns based on actual content length
            # (catches columns the source never explicitly sized)
            from openpyxl.utils import get_column_letter as _gcl
            col_widths = {}
            for row in src_ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        col_letter = _gcl(cell.column)
                        val_str = str(cell.value)
                        cell_len = min(len(val_str), 20)
                        col_widths[col_letter] = max(
                            col_widths.get(col_letter, 6),
                            cell_len + 1,   # tight +1 padding
                        )
            for col_letter, w in col_widths.items():
                # Cap at 22 to keep columns compact
                dest_ws.column_dimensions[col_letter].width = min(w, 22)

            # ── Copy row heights ──────────────────────────────────────
            for row_idx, row_dim in src_ws.row_dimensions.items():
                dest_ws.row_dimensions[row_idx].height = row_dim.height

            # ── Copy merged cells ─────────────────────────────────────
            for merge_range in src_ws.merged_cells.ranges:
                dest_ws.merge_cells(str(merge_range))

            # ── Copy all cell values + number formats ─────────────────
            for row in src_ws.iter_rows():
                for cell in row:
                    dest_cell = dest_ws.cell(
                        row=cell.row,
                        column=cell.column,
                        value=cell.value,
                    )
                    if cell.has_style:
                        dest_cell.number_format = cell.number_format

            # ── Apply formatting per sheet type ───────────────────────
            is_scenario   = _is_scenario_sheet(src_sheet_name)
            is_top6       = _is_top6_sheet(src_sheet_name)
            is_parameters = _is_parameters_sheet(src_sheet_name)

            if is_scenario:
                # Read header row to identify columns
                max_col  = src_ws.max_column
                max_row  = src_ws.max_row
                headers  = {}   # col_idx → header string

                for col_idx in range(1, max_col + 1):
                    hdr_cell = dest_ws.cell(row=1, column=col_idx)
                    val = hdr_cell.value
                    if val is not None:
                        headers[col_idx] = str(val)

                # Style header row
                for col_idx in range(1, max_col + 1):
                    c = dest_ws.cell(row=1, column=col_idx)
                    c.fill      = _make_hdr_fill()
                    c.font      = _make_hdr_font()
                    c.alignment = _center()

                # Style data rows with alternating tint
                for row_idx in range(2, max_row + 1):
                    fill = tint_fill if row_idx % 2 == 0 else _make_white_fill()
                    for col_idx in range(1, max_col + 1):
                        c = dest_ws.cell(row=row_idx, column=col_idx)
                        c.fill      = fill
                        c.font      = _make_data_font()
                        c.alignment = _center()

                # Winner row highlight — find rows matching selected SL/BE (GROSS basis)
                if combined_profiles and profile in combined_profiles:
                    winner_sl = combined_profiles[profile].get("sl")
                    winner_be = combined_profiles[profile].get("be")

                    # Find column indices for SL, BE, Basis
                    sl_col   = next((i for i, h in headers.items() if h == "SL (%)"), None)
                    be_col   = next((i for i, h in headers.items() if h == "BE (after TP #)"), None)
                    bas_col  = next((i for i, h in headers.items() if h == "Basis"), None)

                    if sl_col and be_col and bas_col:
                        winner_fill = PatternFill("solid", fgColor=WINNER_BG[profile])
                        # Same size=10 as data rows — no size increase avoids #### in narrow columns
                        winner_font = Font(
                            name="Calibri", size=10,
                            bold=True, color=WINNER_FG,
                        )
                        for row_idx in range(2, max_row + 1):
                            sl_val  = dest_ws.cell(row=row_idx, column=sl_col).value
                            be_val  = dest_ws.cell(row=row_idx, column=be_col).value
                            bas_val = dest_ws.cell(row=row_idx, column=bas_col).value

                            # Match GROSS rows only (NET row for same combo stays normal)
                            try:
                                sl_match  = sl_val  is not None and abs(float(sl_val)  - float(winner_sl)) < 1e-6
                                be_match  = (
                                    (winner_be is None and (be_val is None or (isinstance(be_val, float) and pd.isna(be_val))))
                                    or (winner_be is not None and be_val is not None and int(float(be_val)) == int(winner_be))
                                )
                                bas_match = str(bas_val).strip().upper() == "GROSS"
                            except Exception:
                                continue

                            if sl_match and be_match and bas_match:
                                for col_idx in range(1, max_col + 1):
                                    c = dest_ws.cell(row=row_idx, column=col_idx)
                                    c.fill = winner_fill
                                    c.font = winner_font
                                # No row height change — keeps all rows uniform

                # Conditional colour scales on data range
                if max_row > 1:
                    for col_idx, hdr in headers.items():
                        col_letter = get_column_letter(col_idx)
                        cf_range   = f"{col_letter}2:{col_letter}{max_row}"

                        if _col_is_good_high(hdr):
                            dest_ws.conditional_formatting.add(
                                cf_range,
                                ColorScaleRule(
                                    start_type="min",  start_color=CF_RED,
                                    mid_type="percentile", mid_value=50,
                                    mid_color=CF_YELLOW,
                                    end_type="max",    end_color=CF_GREEN,
                                )
                            )
                        elif _col_is_good_low(hdr):
                            dest_ws.conditional_formatting.add(
                                cf_range,
                                ColorScaleRule(
                                    start_type="min",  start_color=CF_GREEN,
                                    mid_type="percentile", mid_value=50,
                                    mid_color=CF_YELLOW,
                                    end_type="max",    end_color=CF_RED,
                                )
                            )

            elif is_top6:
                # Top6 sheet: re-open source WITHOUT data_only so we get
                # conditional formatting rules, then copy cells + CF rules.
                try:
                    src_wb_fmt = load_workbook(src_path, data_only=False)
                    src_ws_fmt = src_wb_fmt[src_sheet_name]

                    # Cell values + styles
                    for row in src_ws_fmt.iter_rows():
                        for cell in row:
                            dest_cell = dest_ws.cell(
                                row=cell.row,
                                column=cell.column,
                                value=cell.value,
                            )
                            if cell.has_style:
                                dest_cell.font          = copy(cell.font)
                                dest_cell.fill          = copy(cell.fill)
                                dest_cell.border        = copy(cell.border)
                                dest_cell.alignment     = copy(cell.alignment)
                                dest_cell.number_format = cell.number_format

                    # Copy conditional formatting rules — iterate safely
                    # openpyxl stores CF as ConditionalFormattingList;
                    # iterate over the list directly to avoid internal API changes
                    try:
                        for cf in src_ws_fmt.conditional_formatting:
                            # cf is a ConditionalFormattingList entry:
                            # cf.sqref = range string, cf.rules = list of rules
                            sqref = str(cf.sqref)
                            for rule in cf.rules:
                                dest_ws.conditional_formatting.add(sqref, rule)
                    except Exception as cf_iter_err:
                        print(f"[WARN] Top6 CF iteration skipped: {cf_iter_err}")

                    src_wb_fmt.close()

                except Exception as cf_err:
                    print(f"[WARN] Could not copy Top6 CF rules: {cf_err}")
                    # Fall back to plain style copy from the data_only handle
                    for row in src_ws.iter_rows():
                        for cell in row:
                            dest_cell = dest_ws.cell(row=cell.row, column=cell.column)
                            if cell.has_style:
                                dest_cell.font      = copy(cell.font)
                                dest_cell.fill      = copy(cell.fill)
                                dest_cell.border    = copy(cell.border)
                                dest_cell.alignment = copy(cell.alignment)

            elif is_parameters:
                # Parameters sheet: clean two-column layout, no tinting
                for row in src_ws.iter_rows():
                    for cell in row:
                        dest_cell = dest_ws.cell(row=cell.row, column=cell.column)
                        if cell.has_style:
                            dest_cell.font      = copy(cell.font)
                            dest_cell.fill      = copy(cell.fill)
                            dest_cell.alignment = copy(cell.alignment)

            else:
                # Any other sheet — copy styles as-is
                for row in src_ws.iter_rows():
                    for cell in row:
                        dest_cell = dest_ws.cell(row=cell.row, column=cell.column)
                        if cell.has_style:
                            dest_cell.font      = copy(cell.font)
                            dest_cell.fill      = copy(cell.fill)
                            dest_cell.alignment = copy(cell.alignment)

        src_wb.close()
        print(f"  ✓ Merged {len(src_wb.sheetnames)} sheets from {label}")

    # ------------------------------------------------------------------
    # Embed equity curve PNGs as one combined chart sheet
    # ------------------------------------------------------------------
    if equity_pngs:
        from openpyxl.drawing.image import Image as XLImage

        combined_png = equity_pngs.get("combined")
        if combined_png and Path(combined_png).exists():
            eq_ws = combined_wb.create_sheet(title="📈 Equity Curves")
            eq_ws.sheet_properties.tabColor = "334155"

            dark_fill = PatternFill("solid", fgColor="0F172A")
            for row_idx in range(1, 60):
                for col_idx in range(1, 50):
                    eq_ws.cell(row=row_idx, column=col_idx).fill = dark_fill

            img        = XLImage(str(combined_png))
            img.anchor = "B2"
            eq_ws.add_image(img)
            print(f"  ✓ Combined equity chart embedded → 📈 Equity Curves")

    if not combined_wb.sheetnames:
        print("[WARN] Merge: no sheets were written — combined workbook not saved.")
        return None

    # ------------------------------------------------------------------
    # Build index sheet — inserted as the first tab
    # ------------------------------------------------------------------
    idx_ws = combined_wb.create_sheet(title="📋 Index", index=0)
    idx_ws.sheet_properties.tabColor = "334155"   # dark slate tab

    # Column widths
    idx_ws.column_dimensions["A"].width = 3
    idx_ws.column_dimensions["B"].width = 28
    idx_ws.column_dimensions["C"].width = 18
    idx_ws.column_dimensions["D"].width = 40

    # Palette reused from main formatter
    PROFILE_HEADER_COLORS = {
        "safe":       ("166534", "DCFCE7"),   # dark green text, light green bg
        "balanced":   ("1E40AF", "DBEAFE"),   # dark blue text, light blue bg
        "aggressive": ("991B1B", "FEE2E2"),   # dark red text, light red bg
    }

    title_font  = Font(name="Calibri", size=16, bold=True, color="1E293B")
    sub_font    = Font(name="Calibri", size=10, color="64748B")
    link_font_base = {"name": "Calibri", "size": 11, "underline": "single"}

    # Title
    idx_ws.merge_cells("B2:D2")
    t = idx_ws.cell(row=2, column=2, value="Strategy Backtest Report")
    t.font      = title_font
    t.alignment = Alignment(horizontal="left", vertical="center")

    # Subtitle line
    idx_ws.merge_cells("B3:D3")
    s = idx_ws.cell(row=3, column=2,
                    value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}  |  Window: {selection_window}D")
    s.font      = sub_font
    s.alignment = Alignment(horizontal="left", vertical="center")

    row = 5   # start writing profile blocks from row 5

    # Sheet-type descriptions shown in column D
    SHEET_DESCRIPTIONS = {
        "all trades":    "All trades across full history",
        "last 30d":      "Last 30 days of trades",
        "last 90d":      "Last 90 days of trades",
        "last 180d":     "Last 180 days of trades",
        "last 360d":     "Last 360 days of trades",
        "top6":          "Top 6 SL/BE combos ranked by Profit Factor",
        "parameters":    "Input parameters & TP ladder used for this run",
        "equity curve":  "Safe, Balanced & Aggressive equity curves + drawdown",
    }

    def _sheet_desc(sheet_name: str) -> str:
        n = sheet_name.lower()
        for key, desc in SHEET_DESCRIPTIONS.items():
            if key in n:
                return desc
        return ""

    for profile in profiles:
        label         = PROFILE_LABELS[profile]
        txt_color, bg_color = PROFILE_HEADER_COLORS[profile]

        # Profile header row
        idx_ws.merge_cells(f"B{row}:D{row}")
        hdr = idx_ws.cell(row=row, column=2, value=f"  {label.upper()}")
        hdr.font      = Font(name="Calibri", size=12, bold=True, color=txt_color)
        hdr.fill      = PatternFill("solid", fgColor=bg_color)
        hdr.alignment = Alignment(horizontal="left", vertical="center")
        idx_ws.row_dimensions[row].height = 20
        row += 1

        # Find all sheets belonging to this profile
        profile_sheets = [
            name for name in combined_wb.sheetnames
            if name.lower().startswith(label.lower() + " -")
        ]

        if not profile_sheets:
            idx_ws.cell(row=row, column=2, value="(no sheets found)")
            row += 1
        else:
            for sheet_name in profile_sheets:
                # Strip profile prefix for display: "Safe - Last 30D" → "Last 30D"
                display_name = sheet_name[len(label) + 3:]   # skip "Safe - "

                # Hyperlink cell
                link_cell = idx_ws.cell(row=row, column=2, value=display_name)
                link_cell.hyperlink = f"#'{sheet_name}'!A1"
                link_cell.font      = Font(
                    **link_font_base,
                    color=txt_color,
                )
                link_cell.alignment = Alignment(horizontal="left", vertical="center")

                # Description cell
                desc_cell = idx_ws.cell(row=row, column=4,
                                        value=_sheet_desc(sheet_name))
                desc_cell.font      = sub_font
                desc_cell.alignment = Alignment(horizontal="left", vertical="center")

                idx_ws.row_dimensions[row].height = 18
                row += 1

        row += 1   # blank spacer between profiles

    combined_wb.save(combined_path)
    print(f"\n📊 Combined workbook written → {combined_path.name}")
    return combined_path



# ============================================================
# Signal JSON Export  —  bot-ready format
# ============================================================

def write_signal_jsons(
    base_run_dir: Path,
    ticker: str,
    safe_name: str,
    combined_profiles: dict,
) -> None:
    """
    Write one buy + one sell signal JSON per profile into base_run_dir.

    File naming:  {safe_name}_safe_buy.json  /  {safe_name}_safe_sell.json
                  {safe_name}_balanced_buy.json  ...  etc.

    Format:
      action          "buy" | "sell"
      symbol          {ticker}USDT
      bot_id          "{BOT_ID}"   ← runtime placeholder
      entry           "{ENTRY}"    ← runtime placeholder
      stop_loss_pct   SL value from optimised profile
      stop_loss_atr   "{ATR_SL}"   ← runtime placeholder
      tp1 … tp6       TP levels from optimised profile (%)
      tp7 … tp10      null  (reserved for future expansion to 10 TPs)
    """
    symbol = f"{ticker}USDT"

    signals_dir = base_run_dir / "signals"
    signals_dir.mkdir(parents=True, exist_ok=True)

    for profile, prof_data in combined_profiles.items():
        tps = prof_data.get("tp", [])
        sl  = prof_data.get("sl")

        # Build tp1..tp10 — real values for levels we have, null beyond
        tp_fields = {}
        for i in range(1, 11):
            idx = i - 1
            tp_fields[f"tp{i}"] = round(float(tps[idx]), 2) if idx < len(tps) else None

        for action in ("buy", "sell"):
            signal = {
                "action":        action,
                "symbol":        symbol,
                "bot_id":        "{BOT_ID}",
                "entry":         "{ENTRY}",
                "stop_loss_pct": round(float(sl), 2) if sl is not None else None,
                "stop_loss_atr": "{ATR_SL}",
                **tp_fields,
            }

            filename  = f"{safe_name}_{profile}_{action}.json"
            out_path  = signals_dir / filename

            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(signal, f, indent=2)

            print(f"📤 Signal JSON written → {filename}")



# ============================================================
# Equity Curve Generator
# ============================================================

# Chart colours per profile
EQUITY_COLORS = {
    "safe":       {"line": "#22C55E", "dd": "#16A34A", "title": "Safe"},
    "balanced":   {"line": "#3B82F6", "dd": "#1D4ED8", "title": "Balanced"},
    "aggressive": {"line": "#EF4444", "dd": "#B91C1C", "title": "Aggressive"},
}

def _compute_equity(pnl_series: pd.Series):
    """
    From a series of per-trade PnL % values, compute:
      - cumulative: compounded cumulative return (%)
      - drawdown:   rolling drawdown from peak (%, always <= 0)
    """
    pnl = pnl_series.fillna(0.0).values
    # Compounded equity curve starting at 100
    equity = np.cumprod(1.0 + pnl / 100.0) * 100.0
    # Drawdown from running peak
    peak = np.maximum.accumulate(equity)
    drawdown = (equity / peak - 1.0) * 100.0
    cumulative = equity - 100.0   # shift to 0-based % gain
    return cumulative, drawdown


def _simulate_trade_pnl(
    mfe_pct: float,
    mae_pct: float,
    pnl_pct: float,
    tps: list,
    weights: list,
    sl: float,
    be: int,
) -> float:
    """
    Simulate the blended PnL for one trade given a profile's TP ladder.

    Rules:
      1. If mae_pct <= -sl  → stopped out: return -sl
      2. For each TP level (in order):
           if mfe_pct >= tp  → that weight slice exits at tp
           else              → that weight slice exits at pnl_pct
      3. If be is set and TP{be} was hit → floor remaining slices at 0
         (breakeven protection kicks in)
    """
    if mae_pct <= -abs(sl):
        return -abs(sl)

    be_triggered = False
    blended = 0.0
    remaining_weight = 1.0

    for i, (tp, w) in enumerate(zip(tps, weights)):
        tp_hit = mfe_pct >= tp
        if tp_hit:
            blended += w * tp
            remaining_weight -= w
            # Check if this TP triggers breakeven
            if be is not None and (i + 1) >= int(be):
                be_triggered = True
        else:
            # This TP not hit — remaining slices exit at pnl_pct (or 0 if BE active)
            exit_pnl = max(0.0, pnl_pct) if be_triggered else pnl_pct
            blended += remaining_weight * exit_pnl
            remaining_weight = 0.0
            break

    # Any weight still remaining (all TPs hit) exits at last TP
    if remaining_weight > 1e-9:
        blended += remaining_weight * pnl_pct

    return blended


def generate_equity_curves(
    base_run_dir: Path,
    profiles: list,
    combined_profiles: dict,
    ticker: str,
    selection_window: str,
) -> dict:
    """
    Generate one combined side-by-side equity curve image:
      Safe | Balanced | Aggressive — each column has equity line + drawdown area.
    Returns dict with 'combined' key pointing to the PNG path.
    """
    png_paths  = {}
    chart_data = []

    for profile in profiles:
        trades_path = base_run_dir / profile / "trades.csv"
        if not trades_path.exists():
            print(f"[WARN] Equity curve: trades.csv not found for {profile} — skipping")
            continue

        winner     = combined_profiles.get(profile, {})
        winner_sl  = winner.get("sl")
        winner_be  = winner.get("be")
        winner_tps = winner.get("tp", [])
        winner_w   = winner.get("w", [])
        winner_lev = winner.get("lev", 1) or 1

        if not winner_tps or winner_sl is None:
            print(f"[WARN] Equity curve: missing TP/SL data for {profile} — skipping")
            continue

        try:
            df = pd.read_csv(trades_path)
        except Exception as e:
            print(f"[WARN] Equity curve: could not read trades.csv for {profile}: {e}")
            continue

        if "exit_ts" in df.columns:
            df["_exit_ts"] = pd.to_datetime(df["exit_ts"], utc=True, errors="coerce")
            df = df.dropna(subset=["_exit_ts"])
            window_end   = df["_exit_ts"].max()
            window_start = window_end - pd.Timedelta(days=int(selection_window))
            df = df[df["_exit_ts"] >= window_start].copy()

        if df.empty:
            print(f"[WARN] Equity curve: no trades in window for {profile} — skipping")
            continue

        sim_pnls = df.apply(
            lambda row: _simulate_trade_pnl(
                mfe_pct = float(row["mfe_pct"]),
                mae_pct = float(row["mae_pct"]),
                pnl_pct = float(row["pnl_pct"]),
                tps     = winner_tps,
                weights = winner_w,
                sl      = winner_sl,
                be      = winner_be,
            ),
            axis=1,
        ) * winner_lev

        cumulative, drawdown = _compute_equity(sim_pnls)
        be_label = f"BE TP{int(winner_be)}" if winner_be else "No BE"

        chart_data.append({
            "profile":    profile,
            "cumulative": cumulative,
            "drawdown":   drawdown,
            "trades_x":   np.arange(1, len(cumulative) + 1),
            "colors":     EQUITY_COLORS[profile],
            "title":      (f"{EQUITY_COLORS[profile]['title']}\n"
                           f"SL {winner_sl}%  {be_label}  L{winner_lev}x"),
            "n_trades":   len(cumulative),
            "final":      float(cumulative[-1]),
            "max_dd":     float(drawdown.min()),
        })

    if not chart_data:
        print("[WARN] Equity curve: no data to plot")
        return png_paths

    n_cols = len(chart_data)

    # ── Combined side-by-side image ───────────────────────────────────
    fig = plt.figure(figsize=(7 * n_cols, 9), facecolor="#0F172A")
    fig.suptitle(
        f"{ticker}  ·  Equity Curves  ·  Window {selection_window}D",
        color="#F1F5F9", fontsize=13, fontweight="bold", y=0.98,
    )

    gs = gridspec.GridSpec(
        2, n_cols,
        height_ratios=[3, 1],
        hspace=0.08, wspace=0.12,
        left=0.06, right=0.98,
        top=0.91, bottom=0.07,
    )

    for col_idx, d in enumerate(chart_data):
        ax_eq = fig.add_subplot(gs[0, col_idx])
        ax_dd = fig.add_subplot(gs[1, col_idx], sharex=ax_eq)
        colors     = d["colors"]
        trades_x   = d["trades_x"]
        cumulative = d["cumulative"]
        drawdown   = d["drawdown"]

        for ax in (ax_eq, ax_dd):
            ax.set_facecolor("#1E293B")
            ax.tick_params(colors="#94A3B8", labelsize=8)
            ax.spines[:].set_color("#334155")
            ax.grid(True, color="#334155", linewidth=0.5,
                    linestyle="--", alpha=0.7)

        # Equity line + fill
        ax_eq.plot(trades_x, cumulative, color=colors["line"],
                   linewidth=2.0, zorder=3)
        ax_eq.fill_between(trades_x, 0, cumulative,
                           where=(np.array(cumulative) >= 0),
                           color=colors["line"], alpha=0.12, zorder=2)
        ax_eq.axhline(0, color="#475569", linewidth=0.8)
        ax_eq.set_title(d["title"], color=colors["line"],
                        fontsize=10, fontweight="bold", pad=6)
        ax_eq.annotate(
            f"{d['final']:+.1f}%",
            xy=(trades_x[-1], d["final"]),
            xytext=(-8, 8 if d["final"] >= 0 else -14),
            textcoords="offset points",
            color=colors["line"], fontsize=9, fontweight="bold", ha="right",
        )
        ax_eq.tick_params(labelbottom=False)

        # Y labels only on leftmost column
        if col_idx == 0:
            ax_eq.set_ylabel("Cumulative Return (%)", color="#94A3B8", fontsize=9)
            ax_dd.set_ylabel("Drawdown (%)", color="#94A3B8", fontsize=9)

        # Drawdown area
        ax_dd.fill_between(trades_x, drawdown, 0,
                           color=colors["dd"], alpha=0.5, zorder=2)
        ax_dd.plot(trades_x, drawdown, color=colors["dd"],
                   linewidth=1.2, zorder=3)
        ax_dd.axhline(0, color="#475569", linewidth=0.8)
        ax_dd.set_xlabel("Trade #", color="#94A3B8", fontsize=8)
        ax_dd.annotate(
            f"Max DD: {d['max_dd']:.1f}%",
            xy=(trades_x[int(np.argmin(drawdown))], d["max_dd"]),
            xytext=(6, -4), textcoords="offset points",
            color="#F87171", fontsize=8,
        )
        ax_dd.annotate(
            f"{d['n_trades']} trades",
            xy=(0.02, 0.08), xycoords="axes fraction",
            color="#64748B", fontsize=8,
        )

    combined_png = base_run_dir / "equity_combined.png"
    fig.savefig(combined_png, dpi=150, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    plt.close(fig)
    print(f"📈 Combined equity curve saved → {combined_png.name}")
    png_paths["combined"] = combined_png

    return png_paths



# ============================================================
# Profile worker — must be at module level for Windows
# multiprocessing pickle compatibility
# ============================================================

def _run_profile(profile: str, cfg: dict) -> tuple:
    """
    Run one full profile pipeline (build_trades → strategy_engine → backtest).
    Returns (profile, combined_profile_entry).
    Must be a module-level function so ProcessPoolExecutor can pickle it on Windows.
    """
    import json, sys
    import pandas as pd
    from pathlib import Path

    csv_path         = Path(cfg["csv_path"])
    base_run_dir     = Path(cfg["base_run_dir"])
    selection_window = cfg["selection_window"]
    sl_list          = cfg["sl_list"]
    be_list          = cfg["be_list"]
    maker_fee_pct    = cfg["maker_fee_pct"]
    taker_fee_pct    = cfg["taker_fee_pct"]
    stoploss_type    = cfg["stoploss_type"]
    stoploss_value   = cfg["stoploss_value"]
    now_utc_str      = cfg["now_utc_str"]

    print(f"\n[{profile.upper()}] Starting pipeline...")

    profile_dir = base_run_dir / profile
    profile_dir.mkdir(parents=True, exist_ok=True)

    # STEP 0: Build trades
    build_cmd = [
        sys.executable,
        str(Path("build_trades_from_signals.py").resolve()),
        "--signals", str(csv_path),
        "--out", "trades.csv"
    ]
    run_subprocess(build_cmd, cwd=profile_dir)

    # STEP 0b: Sample trades
    try:
        trades_csv_path = profile_dir / "trades.csv"
        trades_df = pd.read_csv(trades_csv_path)

        trades_df["_exit_ts"] = pd.to_datetime(
            trades_df["exit_ts"], utc=True, errors="coerce"
        )
        trades_df = trades_df.dropna(subset=["_exit_ts"])

        window_end   = trades_df["_exit_ts"].max()
        window_start = window_end - pd.Timedelta(days=int(selection_window))
        window_df    = trades_df[trades_df["_exit_ts"] >= window_start].copy()
        window_df    = window_df.drop(columns=["_exit_ts"])

        n_sample = min(10, len(window_df))

        sample_cols = [
            "entry_date", "entry_time", "side",
            "entry_price", "exit_price",
            "pnl_pct", "mfe_pct", "mae_pct",
        ]
        sample_cols = [c for c in sample_cols if c in window_df.columns]

        sample_df = (
            window_df[sample_cols]
            .sample(n=n_sample)
            .sort_values("entry_date")
            .reset_index(drop=True)
        )

        sample_path = profile_dir / f"sample_trades_{profile}.csv"

        with open(sample_path, "w", encoding="utf-8") as fh:
            fh.write(
                f"# Sample trades — {profile.upper()} | "
                f"window: last {selection_window}D | "
                f"{n_sample} of {len(window_df)} trades | "
                f"generated: {now_utc_str}\n"
            )
            sample_df.to_csv(fh, index=False)

        print(f"[{profile.upper()}] 📋 Sample trades written ({n_sample} trades)")

    except Exception as e:
        print(f"[{profile.upper()}] [WARN] Could not write sample trades: {e}")

    # STEP 1: Strategy engine
    engine_cmd = [
        sys.executable,
        str(Path("strategy_engine.py").resolve()),
        "trades.csv",
        "--lookback_days", selection_window,
        "--profile", profile.upper(),
    ]
    run_subprocess(engine_cmd, cwd=profile_dir)

    strategy_json_path = profile_dir / "strategy_profiles.json"
    if not strategy_json_path.exists():
        raise RuntimeError(f"[{profile.upper()}] Engine did not produce strategy_profiles.json")

    with open(strategy_json_path, "r") as f:
        engine_data = json.load(f)

    engine_profile = engine_data["profiles"][profile]

    # STEP 2: Backtester
    backtest_cmd = [
        sys.executable,
        str(Path("backtest_simulator.py").resolve()),
        "--csv", str(csv_path),
        "--tp", json.dumps(engine_profile["tp"]),
        "--w", json.dumps(engine_profile["w"]),
        "--sl_list", sl_list,
        "--be_list", be_list,
        "--maker_fee_pct", maker_fee_pct,
        "--taker_fee_pct", taker_fee_pct,
        "--table_csv", f"scenario_table_{profile}.csv",
        "--table_xlsx", f"scenario_table_{profile}.xlsx",
        "--stoploss_type", stoploss_type,
        "--stoploss_value", stoploss_value,
        "--selection_window", selection_window,
    ]

    run_subprocess(backtest_cmd, cwd=profile_dir)

    xlsx_path = profile_dir / f"scenario_table_{profile}.xlsx"

    winner = None
    try:
        winner = extract_top6_winner(
            xlsx_path,
            selection_window,
            profile
        )
    except ValueError as e:
        print(f"[{profile.upper()}] [WARN] Could not extract Top6 winner "
              f"({selection_window}D): {e}")
        print(f"[{profile.upper()}] [WARN] Using engine defaults for SL/BE/LEV.")

    if winner is None:
        winner = {
            "sl": float(engine_profile.get("sl", 3.0)),
            "be": engine_profile.get("be", 1),
            "lev": int(engine_profile.get("lev", 2)),
        }

    weights = normalize_weights_exact(engine_profile["w"], decimals=4)

    if round(sum(weights), 6) != 1.0:
        raise ValueError(f"[{profile.upper()}] Weights do not sum to 1.0 after normalization")

    entry = {
        "tp": engine_profile["tp"],
        "w": weights,
        "sl": winner["sl"],
        "be": winner["be"],
        "lev": winner["lev"]
    }

    print(f"[{profile.upper()}] ✅ Pipeline complete.")
    return profile, entry


def main():

    ap = argparse.ArgumentParser()

    ap.add_argument("--csv", required=True)
    ap.add_argument("--selection_window", default="180")
    ap.add_argument("--sl_list", required=True)
    ap.add_argument("--be_list", required=True)
    ap.add_argument("--maker_fee_pct", required=True)
    ap.add_argument("--taker_fee_pct", required=True)
    ap.add_argument("--ticker", required=True)
    ap.add_argument("--exchange", required=True)
    ap.add_argument("--timeframe", required=True)
    ap.add_argument("--name", required=True)
    ap.add_argument("--type", required=True)
    ap.add_argument("--input_baseline", required=True)
    ap.add_argument("--input_sensitivity", required=True)
    ap.add_argument("--stoploss_type", required=True)
    ap.add_argument("--stoploss_value", required=True, type=float)

    args = ap.parse_args()
    now_utc = datetime.now(timezone.utc)
    
    # --------------------------------------------------------
    # Validation
    # --------------------------------------------------------

    try:
        selection_window_int = int(args.selection_window)
        if selection_window_int < 14 or selection_window_int > 730:
            raise ValueError
    except ValueError:
        raise ValueError("selection_window must be a number between 14 and 730")

    try:
        sl_list = json.loads(args.sl_list)
        be_list = json.loads(args.be_list)
    except Exception:
        raise ValueError("sl_list and be_list must be valid JSON lists")

    csv_path = Path(args.csv).resolve()
    if not csv_path.exists():
        raise FileNotFoundError("CSV not found.")

    # --------------------------------------------------------
    # Folder + Naming
    # --------------------------------------------------------

    token = args.ticker.split(".")[0]
    today = datetime.now().strftime("%Y-%m-%d")  # local date, not UTC

    safe_name = re.sub(r"[^A-Za-z0-9_\-]", "_", args.name)
    output_filename = f"{safe_name}.json"

    base_run_dir = Path("runs") / token / today
    base_run_dir.mkdir(parents=True, exist_ok=True)

    # --------------------------------------------------------
    # Save Run Configuration Snapshot
    # --------------------------------------------------------

    config_snapshot = {
        "csv": str(csv_path),
        "selection_window": args.selection_window,
        "sl_list": sl_list,
        "be_list": be_list,
        "maker_fee_pct": float(args.maker_fee_pct),
        "taker_fee_pct": float(args.taker_fee_pct),
        "ticker": args.ticker,
        "exchange": args.exchange,
        "timeframe": args.timeframe,
        "name": args.name,
        "type": args.type,
        "input_baseline": args.input_baseline,
        "input_sensitivity": args.input_sensitivity,
        "stoploss_type": args.stoploss_type,
        "stoploss_value": float(args.stoploss_value),
        "generated_utc": datetime.now(timezone.utc).isoformat()
    }

    config_path = base_run_dir / "run_config.json"

    with open(config_path, "w") as f:
        json.dump(config_snapshot, f, indent=2)

    print(f"Run configuration saved to: {config_path}")

    profiles = ["safe", "balanced", "aggressive"]
    combined_profiles = {}

    # --------------------------------------------------------
    # Per Profile Execution — runs all 3 profiles in parallel
    # --------------------------------------------------------

    # Build a plain serialisable config dict — needed because ProcessPoolExecutor
    # on Windows can only pickle module-level functions and simple data types.
    profile_cfg = {
        "csv_path":         str(csv_path),
        "base_run_dir":     str(base_run_dir),
        "selection_window": args.selection_window,
        "sl_list":          args.sl_list,
        "be_list":          args.be_list,
        "maker_fee_pct":    args.maker_fee_pct,
        "taker_fee_pct":    args.taker_fee_pct,
        "stoploss_type":    args.stoploss_type,
        "stoploss_value":   str(args.stoploss_value),
        "now_utc_str":      now_utc.strftime("%Y-%m-%d %H:%M UTC"),
    }

    from concurrent.futures import ProcessPoolExecutor

    print("\n⚡ Running Safe / Balanced / Aggressive profiles in parallel...")

    futures_map = {}
    with ProcessPoolExecutor(max_workers=3) as pool:
        for profile in profiles:
            future = pool.submit(_run_profile, profile, profile_cfg)
            futures_map[future] = profile

    errors = []
    for future, profile in futures_map.items():
        try:
            prof_name, prof_entry = future.result()
            combined_profiles[prof_name] = prof_entry
        except Exception as exc:
            errors.append((profile, exc))
            print(f"❌ Profile '{profile}' failed: {exc}")

    if errors:
        print("❌ One or more profiles failed — aborting.")
        sys.exit(1)

    # --------------------------------------------------------
    # Build Combined Strategy JSON
    # --------------------------------------------------------

    if not combined_profiles:
        print("❌ All profiles failed — combined_profiles is empty. "
              "Aborting JSON write to prevent overwriting active strategy.")
        sys.exit(1)

    # Split baseline into name + numeric value
    baseline_parts = args.input_baseline.strip().split()

    if len(baseline_parts) >= 2:
        baseline_name = baseline_parts[0]
        try:
            baseline_value = float(baseline_parts[1])
        except ValueError:
            baseline_value = None
    else:
        baseline_name = args.input_baseline
        baseline_value = None

    combined_strategy = {
        "name": args.name,
        "ticker": args.ticker,
        "optimized_period": int(args.selection_window),
        "exchange": args.exchange,
        "timeframe": args.timeframe,
        "baseline": baseline_name,
        "baseline_value": baseline_value,
        "signal_setting": float(args.input_sensitivity),
        "stoploss_type": args.stoploss_type,
        "stoploss_value": float(args.stoploss_value),
        "type": args.type,
        "fees": {
            "maker_fee_pct": float(args.maker_fee_pct),
            "taker_fee_pct": float(args.taker_fee_pct)
        },
        "engine_version": "v1.0.0",
        "generated_utc": now_utc.strftime("%Y-%m-%d %H:%M:%S UTC"),
        "profiles": combined_profiles
    }

    combined_path = base_run_dir / output_filename

    with open(combined_path, "w") as f:
        json.dump(combined_strategy, f, indent=2)

    # --------------------------------------------------------
    # Write bot signal JSONs (buy + sell per profile)
    # --------------------------------------------------------

    try:
        write_signal_jsons(
            base_run_dir,
            args.ticker,
            safe_name,
            combined_profiles,
        )
    except Exception as e:
        print(f"[WARN] Could not write signal JSONs: {e}")

    # --------------------------------------------------------
    # Copy Replay Artifacts Into Dated Folder
    # --------------------------------------------------------

    try:
        # Copy replay equity json if exists
        replay_equity_path = combined_path.with_suffix(".replay_equity.json")

        if replay_equity_path.exists():
            shutil.copy(replay_equity_path, base_run_dir / replay_equity_path.name)

        # Search for PNGs anywhere inside dated run folder
        for png_path in base_run_dir.rglob("*_equity_curve.png"):
            target_path = base_run_dir / png_path.name

            if png_path.resolve() != target_path.resolve():
                shutil.copy(png_path, target_path)

        print("📦 Replay artifacts copied to dated folder.")

    except Exception as e:
        print(f"[WARN] Could not copy replay artifacts: {e}")

    # --------------------------------------------------------
    # Active + Archive
    # --------------------------------------------------------

    active_dir = Path("runs") / token / "active"
    archive_dir = Path("runs") / token / "archive"

    active_dir.mkdir(parents=True, exist_ok=True)
    archive_dir.mkdir(parents=True, exist_ok=True)

    active_path = active_dir / output_filename

    old_strategy = None
    if active_path.exists():
        with open(active_path, "r") as f:
            old_strategy = json.load(f)

    should_update, gate_reason = stability_gate(old_strategy, combined_strategy)

    if should_update:
        print(f"⚡ Strategy update triggered: {gate_reason}")

        if active_path.exists():
            archive_name = f"{safe_name}_{today}.json"
            shutil.move(str(active_path), archive_dir / archive_name)

        tmp_path = active_dir / (output_filename + ".tmp")
        shutil.copy(combined_path, tmp_path)
        tmp_path.replace(active_path)

    else:
        print(f"✔ Stability gate held: {gate_reason}")
        print("  Active strategy unchanged — no update pushed to bots.")

    # --------------------------------------------------------
    # Inject stability gate result into HTML reports
    # --------------------------------------------------------

    if should_update:
        gate_color  = "#22c55e"
        gate_label  = "UPDATED"
        gate_icon   = "⚡"
    else:
        gate_color  = "#f59e0b"
        gate_label  = "HELD — no bot update"
        gate_icon   = "✔"

    prev_generated = (
        old_strategy.get("generated_utc", "unknown")
        if old_strategy else "N/A — first run"
    )

    gate_html = f"""
<div style='background:#111827;border:1px solid #1f2937;border-radius:10px;
padding:16px;margin-bottom:24px;border-left:4px solid {gate_color}'>
<h2 style='color:{gate_color};margin-top:0;margin-bottom:8px;font-size:16px'>
{gate_icon} Stability Gate: {gate_label}</h2>
<p style='font-size:13px;color:#e5e7eb;margin:4px 0'>
<b>Reason:</b> {gate_reason}</p>
<p style='font-size:12px;color:#9ca3af;margin:4px 0'>
<b>Previous strategy generated:</b> {prev_generated}</p>
<p style='font-size:12px;color:#9ca3af;margin:4px 0'>
<b>Thresholds:</b> TP levels &gt;10% relative move &nbsp;|&nbsp;
Weights &gt;4pp absolute &nbsp;|&nbsp; Any SL / BE / LEV change</p>
</div>"""

    for profile in profiles:
        profile_dir = base_run_dir / profile
        # Engine names the report after the trades csv
        html_report_path = profile_dir / "trades.strategy_report.html"
        if html_report_path.exists():
            try:
                html_content = html_report_path.read_text(encoding="utf-8")
                html_content = html_content.replace(
                    "<div class='container'>",
                    f"<div class='container'>{gate_html}",
                    1
                )
                html_report_path.write_text(html_content, encoding="utf-8")
            except Exception as e:
                print(f"[WARN] Could not inject gate info into {profile} HTML: {e}")

    # --------------------------------------------------------
    # Generate equity curves (PNG + embed into combined xlsx)
    # --------------------------------------------------------

    try:
        equity_pngs = generate_equity_curves(
            base_run_dir,
            profiles,
            combined_profiles,
            args.ticker,
            args.selection_window,
        )
    except Exception as e:
        print(f"[WARN] Could not generate equity curves: {e}")
        equity_pngs = {}

    # --------------------------------------------------------
    # Merge per-profile workbooks into one combined xlsx
    # --------------------------------------------------------

    try:
        merged_path = merge_profile_workbooks(
            base_run_dir,
            profiles,
            safe_name,
            args.selection_window,
            combined_profiles,
            equity_pngs,
        )
    except Exception as e:
        print(f"[WARN] Could not create combined workbook: {e}")
        merged_path = None

    # --------------------------------------------------------
    # Summary
    # --------------------------------------------------------

    print("\n===================================")
    print("RUN COMPLETED SUCCESSFULLY")
    print(f"Token: {args.ticker}")
    print(f"Window: {args.selection_window}D")
    print(f"Run folder: {base_run_dir}")
    print(f"Active file: {active_path}")
    if merged_path:
        print(f"Combined xlsx: {merged_path}")
    print("===================================")


if __name__ == "__main__":
    main()